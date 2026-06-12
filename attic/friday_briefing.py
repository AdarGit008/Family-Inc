"""
Family inc. — Friday Briefing (Phase 2)

Strava-meets-Morning-Brew Friday recap: (1) week spend vs 4w avg,
(2) kids' biggest moment, (3) next week's 3 reminders, (4) one goal
nudge (round-robin), (5) one contract heads-up, plus Shabbat times.

Live mode reads Family_OS.xlsx; falls back to mock data (banner shown).
Claude API used for prose if ANTHROPIC_API_KEY is set, else template.
Cost target <= $0.10/run.

Run: python friday_briefing.py [--dry-run] [--as-of YYYY-MM-DD]
"""
from __future__ import annotations
import argparse, logging, os
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import hebcal_client

ROOT = Path(__file__).parent
SHEET_PATH = ROOT.parent / "Family_OS.xlsx"
BRIEFINGS_DIR = ROOT.parent / "Briefings"

log = logging.getLogger("friday")

# Mock data fallback (matches sunday_briefing.py convention)
MOCK = {
    "week_spend": 3120, "avg_4w_spend": 2890,
    "kids_moment": {"child": "Kid A", "date": "2026-05-28",
                    "event": "First swim class — floated on his back for 5 sec"},
    "next_week_reminders": [
        {"title": "Baby — 4mo Tipat Halav", "date": "2026-06-02", "domain": "Health"},
        {"title": "Mortgage advisor call", "date": "2026-06-03", "domain": "Finance"},
        {"title": "Car insurance renewal", "date": "2026-06-05", "domain": "Car"},
    ],
    "goals": [
        {"name": "House — ⟨town⟩", "owner": "Both", "delta_week": 4200, "last_update_days": 2, "target_year": 2028, "goal_type": "savings"},
        {"name": "Adar — job + grad school", "owner": "Adar", "delta_week": 0, "last_update_days": 24, "target_year": 2027, "goal_type": "career"},
        {"name": "Passive income stream", "owner": "Adar", "delta_week": 180, "last_update_days": 9, "target_year": 2029, "goal_type": "income"},
        {"name": "Shanee — grooming/wellness habit", "owner": "Shanee", "delta_week": 1, "last_update_days": 5, "target_year": 2026, "goal_type": "habit"},
    ],
    "contracts": [
        {"name": "Cellcom mobile family plan", "renews": "2026-07-15", "monthly_ils": 220},
        {"name": "Yes TV bundle", "renews": "2026-08-30", "monthly_ils": 189},
    ],
}

def gather(today: date) -> dict:
    """Return data bag for the briefing. Tries to read live data from Family_OS.xlsx.
    Falls back to MOCK data (with banner) if the sheet is missing or loading fails."""
    data = dict(MOCK)
    try:
        from openpyxl import load_workbook
        if not SHEET_PATH.exists():
            raise FileNotFoundError(f"{SHEET_PATH} not found")
        wb = load_workbook(SHEET_PATH, data_only=True)
        _overlay_live_data(wb, data, today)
        data["_mock"] = False
        return data
    except Exception as e:
        log.warning("sheet load failed: %s — running in mock mode", e)
    data["_mock"] = True
    return data


def _overlay_live_data(wb, data: dict, today: date) -> None:
    """Overwrite mock keys with live values from the workbook where data is present."""
    from datetime import timedelta

    # --- Finance: week spend + 4-week average from Finance-Txns ---
    if "Finance-Txns" in wb.sheetnames:
        ws = wb["Finance-Txns"]
        week_start = today - timedelta(days=today.weekday())  # Monday
        week_amounts: list[float] = []
        monthly_totals: dict[tuple, float] = {}  # (year, week_of_year) -> total
        for row in ws.iter_rows(min_row=2, values_only=True):
            raw_date, _acct, _desc, amount = row[0], row[1], row[2], row[3]
            if raw_date is None or amount is None:
                continue
            try:
                if isinstance(raw_date, datetime):
                    txn_date = raw_date.date()
                elif isinstance(raw_date, date):
                    txn_date = raw_date
                else:
                    txn_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
                amt = float(amount)
            except (ValueError, TypeError):
                continue
            if amt >= 0:  # skip income/credits
                continue
            abs_amt = abs(amt)
            if txn_date >= week_start:
                week_amounts.append(abs_amt)
            # Bin into ISO weeks for 4-week avg
            yr, wk, _ = txn_date.isocalendar()
            monthly_totals[(yr, wk)] = monthly_totals.get((yr, wk), 0.0) + abs_amt
        if week_amounts:
            data["week_spend"] = sum(week_amounts)
        # 4-week average: use the 4 most recent complete weeks (exclude current)
        cur_yr, cur_wk, _ = today.isocalendar()
        past_weeks = [(k, v) for k, v in monthly_totals.items() if k != (cur_yr, cur_wk)]
        past_weeks.sort(reverse=True)
        if past_weeks:
            data["avg_4w_spend"] = sum(v for _, v in past_weeks[:4]) / min(4, len(past_weeks))

    # --- Kids: most recent Education event as the "kids moment" ---
    if "Education" in wb.sheetnames:
        ws = wb["Education"]
        upcoming: list[tuple[date, dict]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            child, _inst, grade, _teacher, raw_date, etype, action = (row[i] for i in range(7))
            if child is None or raw_date is None:
                continue
            try:
                if isinstance(raw_date, datetime):
                    evt_date = raw_date.date()
                elif isinstance(raw_date, date):
                    evt_date = raw_date
                else:
                    evt_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            if evt_date >= today:
                upcoming.append((evt_date, {
                    "child": str(child).strip("[]"),
                    "date": evt_date.isoformat(),
                    "event": str(etype or action or grade or "upcoming event"),
                }))
        if upcoming:
            upcoming.sort()
            data["kids_moment"] = upcoming[0][1]

    # --- Reminders: next 3 upcoming from Reminders tab ---
    if "Reminders" in wb.sheetnames:
        ws = wb["Reminders"]
        fires: list[tuple[date, dict]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            title, domain, _owner, raw_due = row[0], row[1], row[2], row[3]
            if title is None or raw_due is None:
                continue
            if str(title).startswith("["):
                continue
            try:
                if isinstance(raw_due, datetime):
                    due = raw_due.date()
                elif isinstance(raw_due, date):
                    due = raw_due
                else:
                    due = datetime.strptime(str(raw_due)[:10], "%Y-%m-%d").date()
            except (ValueError, TypeError):
                continue
            if due >= today:
                fires.append((due, {
                    "title": str(title),
                    "date": due.isoformat(),
                    "domain": str(domain or "Other"),
                }))
        if fires:
            fires.sort()
            data["next_week_reminders"] = [r for _, r in fires[:3]]

    # --- Goals: read live from Goals tab ---
    if "Goals" in wb.sheetnames:
        ws = wb["Goals"]
        goals = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, owner, _horizon, target_dt, _milestone, pct, last_update_dt, status = (
                row[i] for i in range(8))
            if name is None:
                continue
            if str(name).startswith("["):
                continue
            try:
                target_year = target_dt.year if isinstance(target_dt, (date, datetime)) else int(str(target_dt)[:4])
            except (TypeError, ValueError):
                target_year = today.year + 2
            last_update_days = 999
            if last_update_dt is not None:
                try:
                    if isinstance(last_update_dt, datetime):
                        lu = last_update_dt.date()
                    elif isinstance(last_update_dt, date):
                        lu = last_update_dt
                    else:
                        lu = datetime.strptime(str(last_update_dt)[:10], "%Y-%m-%d").date()
                    last_update_days = (today - lu).days
                except (ValueError, TypeError):
                    pass
            # delta_week not tracked in sheet yet — default 0 until write-back is live
            goals.append({
                "name": str(name),
                "owner": str(owner or "Both"),
                "delta_week": 0,
                "last_update_days": last_update_days,
                "target_year": target_year,
                "goal_type": "",  # not yet in sheet; set via manual data entry
            })
        if goals:
            data["goals"] = goals

    # --- Contracts: read live from Contracts tab ---
    if "Contracts" in wb.sheetnames:
        ws = wb["Contracts"]
        contracts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, _provider, _ctype, _start, renewal_dt, monthly_cost = (row[i] for i in range(6))
            if name is None:
                continue
            if str(name).startswith("["):
                continue
            renews_str = None
            try:
                if isinstance(renewal_dt, datetime):
                    renews_str = renewal_dt.date().isoformat()
                elif isinstance(renewal_dt, date):
                    renews_str = renewal_dt.isoformat()
                elif renewal_dt and str(renewal_dt).strip() not in ("—", ""):
                    renews_str = str(renewal_dt)[:10]
            except (TypeError, ValueError):
                pass
            if renews_str:
                contracts.append({
                    "name": str(name).strip("[]"),
                    "renews": renews_str,
                    "monthly_ils": float(monthly_cost) if monthly_cost else 0,
                })
        if contracts:
            data["contracts"] = contracts

def fmt_money(n) -> str:
    if n is None: return "—"
    try: n = float(n)
    except (TypeError, ValueError): return str(n)
    return f"₪{n:,.0f}"

def section_spend(data: dict) -> str:
    spend, avg = data["week_spend"], data["avg_4w_spend"]
    delta = spend - avg
    pct = (delta / avg * 100) if avg else 0
    arrow = "up" if delta > 0 else ("down" if delta < 0 else "flat")
    return (f"This week: {fmt_money(spend)} spent ({arrow} {abs(pct):.0f}% vs 4-week avg "
            f"of {fmt_money(avg)}).")

def section_kids(data: dict) -> str:
    m = data["kids_moment"]
    return f"Biggest kid moment — {m['child']} on {m['date']}: {m['event']}."

def section_next_week(data: dict) -> str:
    rs = data["next_week_reminders"][:3]
    if not rs:
        return "Next week: clean slate."
    lines = ["Looking ahead to next week:"]
    for r in rs:
        lines.append(f"  {r['date']} — {r['title']} [{r['domain']}]")
    return "\n".join(lines)

def _goal_index(today: date) -> int:
    """Round-robin by ISO week so all four goals cycle monthly."""
    iso_year, iso_week, _ = today.isocalendar()
    return iso_week % 4

def section_goal_nudge(data: dict, today: date) -> str:
    g = data["goals"][_goal_index(today) % len(data["goals"])]
    name, d, days = g["name"], g["delta_week"], g["last_update_days"]
    goal_type = g.get("goal_type", "")  # data-driven: "savings", "habit", "career", etc.
    if d > 0 and days <= 7:
        if goal_type == "savings":
            return f"Goal nudge — {name}: +{fmt_money(d)} this week, on track for {g['target_year']}."
        if goal_type == "habit":
            return f"Goal nudge — {name}: {d} session(s) logged this week. Streak alive."
        return f"Goal nudge — {name}: +{d} this week."
    if days >= 21:
        return f"Goal nudge — {name}: no movement in {days} days. Want to add one task?"
    return f"Goal nudge — {name}: steady, last update {days}d ago."

def section_contract(data: dict, today: date) -> str:
    soon = []
    for c in data["contracts"]:
        try: d = datetime.strptime(c["renews"], "%Y-%m-%d").date()
        except ValueError: continue
        if 0 <= (d - today).days <= 60:
            soon.append(((d - today).days, c))
    if not soon: return "No contract renewals in the next 60 days."
    soon.sort(); days, c = soon[0]
    return (f"Contract heads-up — {c['name']} renews in {days} days "
            f"({c['renews']}, ~{fmt_money(c['monthly_ils'])}/mo). Worth a price-check call?")

def section_shabbat(today: date) -> str:
    st = hebcal_client.shabbat_times(today, next=True)
    if st.get("_stub"):
        return "Shabbat times unavailable (hebcal offline)."
    cl = (st.get("candle_lighting") or "")[11:16] or "?"
    hv = (st.get("havdalah") or "")[11:16] or "?"
    par = st.get("parasha") or "?"
    return f"Shabbat — candles {cl}, havdalah {hv}. Parashat {par}. Shabbat shalom."

# Renderers — templated + (optional) Claude
def render_template(data: dict, today: date) -> str:
    banner = "_RUNNING IN MOCK MODE_\n\n" if data.get("_mock") else ""
    lines = [
        f"# {today.strftime('%A')} family briefing",
        f"_{today.strftime('%B %-d, %Y')}_",
        "",
    ]
    if banner.strip():
        lines.append(banner.strip())
        lines.append("")

    # Money — wrap spend detail in a conversational sentence.
    lines.append(section_spend(data))
    lines.append("")

    # Kids — highlight the biggest moment of the week.
    lines.append(section_kids(data))
    lines.append("")

    # Next week's three things.
    lines.append(section_next_week(data))
    lines.append("")

    # Goal nudge — one goal, round-robin.
    lines.append(section_goal_nudge(data, today))
    lines.append("")

    # Contracts — heads-up if renewals are near.
    lines.append(section_contract(data, today))
    lines.append("")

    # Shabbat times.
    lines.append(section_shabbat(today))
    lines.append("")
    lines.append("---")
    lines.append("_Auto-generated. Edits go back into Family_OS.xlsx._")

    return "\n".join(lines) + "\n"

def render_with_claude(data: dict, today: date) -> Optional[str]:
    """Rewrite via Anthropic SDK; return None if key/SDK missing so caller falls back."""
    key = os.environ.get("ANTHROPIC_API_KEY")
    if not key: return None
    try: import anthropic
    except ImportError:
        log.warning("anthropic SDK not installed — falling back to template"); return None
    raw = render_template(data, today)
    prompt = ("Rewrite this Friday family briefing as a 6-paragraph Morning-Brew-style note: "
              "warm, concrete, no emojis, Hebrew place names kept as-is. "
              "Preserve every number and date. <=220 words total.\n\n" + raw)
    try:  # Haiku keeps cost well under $0.10
        msg = anthropic.Anthropic(api_key=key).messages.create(
            model="claude-haiku-4-5", max_tokens=600,
            messages=[{"role": "user", "content": prompt}])
        return msg.content[0].text
    except Exception as e:
        log.warning("Claude synthesis failed (%s) — using template", e); return None

# Main
def run(today: date, dry_run: bool = False) -> Optional[Path]:
    data = gather(today)
    body = render_with_claude(data, today) or render_template(data, today)
    if dry_run:
        print(body)
        return None
    BRIEFINGS_DIR.mkdir(exist_ok=True)
    out = BRIEFINGS_DIR / f"friday_{today.isoformat()}.md"
    out.write_text(body, encoding="utf-8")
    print(f"wrote {out}")
    print(body)
    return out

def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser()
    ap.add_argument("--as-of", help="YYYY-MM-DD; defaults to today")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    today = datetime.strptime(args.as_of, "%Y-%m-%d").date() if args.as_of else date.today()
    run(today, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
