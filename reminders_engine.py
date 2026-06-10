"""
Family inc. — Reminders Engine (Phase 2, dry-run)

Reads the Reminders tab of Family_OS.xlsx, applies the lead-time and
escalation logic from 02_Reminders_Engine_Spec.md, and writes:

  - /Briefings/{YYYY-MM-DD}_briefing.md     (one digest per recipient)
  - /Briefings/reminders_log.csv            (append-only audit trail)

Port target (decision 2026-06-04, Baileys-first): Automation/wa_outbox.queue_message()
swaps in for write_briefing(). No Twilio.
When the Sheet moves to Google Drive, read_reminders() swaps in for openpyxl.

Run modes:
  python reminders_engine.py            # uses today's date, writes files
  python reminders_engine.py --dry-run  # prints to stdout, writes nothing
  python reminders_engine.py --as-of 2026-06-15  # simulate any date
"""
from __future__ import annotations
import argparse
import csv
import os
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

# ---------------------------------------------------------------------------
# Config — see "From config" in 02_Reminders_Engine_Spec.md
# ---------------------------------------------------------------------------
ROOT = Path(__file__).parent
SHEET_PATH = ROOT / "Family_OS.xlsx"
BRIEFINGS_DIR = ROOT / "Briefings"
LOG_PATH = BRIEFINGS_DIR / "reminders_log.csv"

ALERT_BUDGET_PER_DAY = 2
OVERDUE_REPEAT_DAYS = 3
TOMBSTONE_SKIP_HOURS = 6   # Phase 6.1 offline-queue race guard (see spec §"The daily run")
DROP_FIRST_DOMAINS = {"Goals"}   # de-prioritised — covered by Friday report
ALWAYS_INCLUDE_DOMAINS = {"Health"}  # never trimmed

FLAG_EMOJI = {
    "OVERDUE":    "🔴",
    "FIRE TODAY": "🟠",
    "WEEK OUT":   "🟡",
    "MONTH OUT":  "🟢",
}

# Owner → recipient identity. Port target: logical 'adar'/'shanee' for wa_outbox
# (numbers live only in the bridge machine's recipients.json).
RECIPIENTS = {
    "Adar":    {"name": "Adar",    "channel": "email", "address": "user@example.com"},
    "Partner": {"name": "Partner", "channel": "email", "address": "partner@example.com"},
    "Both":    None,  # expanded at fire time to Adar + Partner
}

# ---------------------------------------------------------------------------
# Data model
# ---------------------------------------------------------------------------
@dataclass
class Reminder:
    row: int
    title: str
    domain: str
    owner: str
    due: date | None
    lead_times: list[int]
    recurrence: str
    status: str
    last_sent: date | None
    channel: str
    notes: str
    # Phase 6.1 — dashboard write-back columns (see 02_Reminders_Engine_Spec.md)
    last_done_by: str = ""
    done_at: datetime | None = None
    write_queue_tombstone: datetime | None = None

@dataclass
class Fire:
    reminder: Reminder
    reason: str       # OVERDUE / FIRE TODAY / WEEK OUT / MONTH OUT / LEAD-TIME / DUE TODAY
    days_until: int

@dataclass
class Digest:
    recipient: str
    fires: list[Fire] = field(default_factory=list)
    dropped: list[Fire] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Read layer (will be swapped for Google Sheets API in production)
# ---------------------------------------------------------------------------
def parse_lead_times(raw) -> list[int]:
    if raw is None: return [7, 1]
    if isinstance(raw, (int, float)): return [int(raw)]
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    out = []
    for p in parts:
        try: out.append(int(p))
        except ValueError: pass
    return sorted(out, reverse=True) or [7, 1]

def to_date(v):
    if v is None: return None
    if isinstance(v, date) and not isinstance(v, datetime): return v
    if isinstance(v, datetime): return v.date()
    if isinstance(v, str):
        try: return datetime.strptime(v.strip(), "%Y-%m-%d").date()
        except ValueError: return None
    return None

def to_datetime(v) -> datetime | None:
    """Best-effort ISO datetime parser. Handles datetime cells, date cells,
    and a few common ISO-ish string shapes the dashboard write-back emits."""
    if v is None: return None
    if isinstance(v, datetime): return v
    if isinstance(v, date): return datetime(v.year, v.month, v.day)
    if isinstance(v, str):
        s = v.strip()
        if not s: return None
        # Tolerate trailing Z and space-separator
        s = s.replace("Z", "+00:00").replace(" ", "T", 1) if "T" not in s and " " in s else s.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(s)
            # Engine compares against naive local time (Asia/Jerusalem). Drop tz if present.
            if dt.tzinfo is not None:
                dt = dt.astimezone().replace(tzinfo=None)
            return dt
        except ValueError:
            try:
                return datetime.strptime(s, "%Y-%m-%d")
            except ValueError:
                return None
    return None

def read_reminders(path: Path) -> list[Reminder]:
    wb = load_workbook(path, data_only=True)
    ws = wb["Reminders"]
    out = []
    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row_idx, 1).value
        if not title or str(title).startswith("["):  # skip blanks / templated rows
            continue
        out.append(Reminder(
            row=row_idx,
            title=str(title),
            domain=str(ws.cell(row_idx, 2).value or "Other"),
            owner=str(ws.cell(row_idx, 3).value or "Adar"),
            due=to_date(ws.cell(row_idx, 4).value),
            lead_times=parse_lead_times(ws.cell(row_idx, 5).value),
            recurrence=str(ws.cell(row_idx, 6).value or "One-off"),
            status=str(ws.cell(row_idx, 7).value or "Pending"),
            last_sent=to_date(ws.cell(row_idx, 8).value),
            channel=str(ws.cell(row_idx, 9).value or "WhatsApp"),
            notes=str(ws.cell(row_idx, 10).value or "").strip(),
            # Phase 6.1 columns M / N / O — blank on legacy sheets, that's fine.
            last_done_by=str(ws.cell(row_idx, 13).value or "").strip(),
            done_at=to_datetime(ws.cell(row_idx, 14).value),
            write_queue_tombstone=to_datetime(ws.cell(row_idx, 15).value),
        ))
    return out


# ---------------------------------------------------------------------------
# Phase 6.1 — tombstone guard (offline-queue race window)
# ---------------------------------------------------------------------------
def is_tombstoned(r: Reminder, now: datetime, window_hours: int = TOMBSTONE_SKIP_HOURS) -> bool:
    """True if the dashboard wrote (or flushed a queued write) within the last
    `window_hours`. When True, the engine MUST skip this row — the sheet state
    may be one hop behind reality. See spec §"The daily run" step 2a."""
    if r.write_queue_tombstone is None:
        return False
    age = now - r.write_queue_tombstone
    if age.total_seconds() < 0:
        # Future-dated tombstone — treat as fresh, not as expired.
        return True
    return age < timedelta(hours=window_hours)


# ---------------------------------------------------------------------------
# Decide-what-fires
# ---------------------------------------------------------------------------
def classify(r: Reminder, today: date) -> Fire | None:
    if r.status in {"Done", "Skipped"}: return None
    if r.due is None: return None
    days = (r.due - today).days

    # Overdue cooldown
    if days < 0:
        if r.last_sent and (today - r.last_sent).days < OVERDUE_REPEAT_DAYS:
            return None
        return Fire(r, "OVERDUE", days)

    if days == 0:
        return Fire(r, "FIRE TODAY", 0)
    if days in r.lead_times:
        if days <= 1:   return Fire(r, "FIRE TODAY", days)
        if days <= 7:   return Fire(r, "WEEK OUT", days)
        if days <= 30:  return Fire(r, "MONTH OUT", days)
        return Fire(r, f"LEAD-{days}", days)
    return None


# ---------------------------------------------------------------------------
# Route to recipients
# ---------------------------------------------------------------------------
def route(fires: list[Fire]) -> dict[str, Digest]:
    digests: dict[str, Digest] = {}
    for f in fires:
        recipients = []
        if f.reminder.owner == "Both":
            recipients = ["Adar", "Partner"]
        elif f.reminder.owner in RECIPIENTS:
            recipients = [f.reminder.owner]
        else:
            recipients = ["Adar"]  # default fallback
        for r in recipients:
            digests.setdefault(r, Digest(recipient=r)).fires.append(f)
    return digests


# ---------------------------------------------------------------------------
# Alert budget (priority-aware trimming)
# ---------------------------------------------------------------------------
PRIORITY = {"OVERDUE": 0, "FIRE TODAY": 1, "WEEK OUT": 2, "MONTH OUT": 3}

def apply_budget(d: Digest, budget: int = ALERT_BUDGET_PER_DAY) -> None:
    # Each digest counts as ONE message regardless of items, so the budget
    # really only bites if more than `budget` digests exist for the same
    # recipient in one run. In single-digest mode we still trim items to keep
    # the message short (top 5).
    must = [f for f in d.fires if f.reason == "OVERDUE"
            or f.reason == "FIRE TODAY"
            or f.reminder.domain in ALWAYS_INCLUDE_DOMAINS]
    rest = [f for f in d.fires if f not in must]
    rest.sort(key=lambda f: (PRIORITY.get(f.reason, 9),
                             f.reminder.domain in DROP_FIRST_DOMAINS,
                             f.days_until))
    keep = must + rest
    if len(keep) > 5:
        d.dropped = keep[5:]
        keep = keep[:5]
    d.fires = keep


# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------
def due_phrase(f: Fire) -> str:
    if f.days_until < 0:  return f"overdue by {-f.days_until} day{'s' if f.days_until != -1 else ''}"
    if f.days_until == 0: return "due today"
    return f"due in {f.days_until} days ({f.reminder.due.isoformat()})"

def render_digest(d: Digest, today: date) -> str:
    head = f"🏠 Family inc. — {today.isoformat()}"
    if not d.fires:
        return f"{head}\n(no reminders today — quiet day.)"
    if len(d.fires) == 1:
        f = d.fires[0]
        emoji = FLAG_EMOJI.get(f.reason, "•")
        body = f"{head}\n{emoji} {f.reminder.title}  ·  {due_phrase(f)}"
        if f.reminder.notes:
            body += f"\n{f.reminder.notes}"
        body += "\n\nReply:  ✅ done    📆 +N days    🤐 mute 30d"
        return body
    lines = [head, f"You have {len(d.fires)} reminders today:"]
    for i, f in enumerate(d.fires, 1):
        emoji = FLAG_EMOJI.get(f.reason, "•")
        lines.append(f"{i}. {emoji} {f.reminder.title} — {due_phrase(f)}")
    if d.dropped:
        lines.append(f"\n(+{len(d.dropped)} more in the dashboard)")
    lines.append("\nReply N ✅ to mark done, or N +D to snooze D days.")
    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Persist
# ---------------------------------------------------------------------------
def write_briefing(d: Digest, today: date, briefings_dir: Path) -> Path:
    briefings_dir.mkdir(exist_ok=True)
    fname = f"{today.isoformat()}_briefing_{d.recipient.lower()}.md"
    path = briefings_dir / fname
    body = render_digest(d, today)
    path.write_text(body + "\n", encoding="utf-8")
    return path

LOG_HEADER = [
    "run_date", "recipient", "fires_sent", "fires_dropped",
    "skipped_due_to_tombstone", "dry_run", "titles_sent",
]

def append_log(today: date, digests: dict[str, Digest], log_path: Path,
               skipped_tombstone: int = 0, dry_run: bool = False) -> None:
    """One row per recipient per run. `skipped_due_to_tombstone` is the per-run
    count and is repeated on each recipient row (it's a run-level metric)."""
    log_path.parent.mkdir(exist_ok=True)
    new_file = not log_path.exists()
    with log_path.open("a", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        if new_file:
            w.writerow(LOG_HEADER)
        # If there are no digests but the run still happened (all tombstoned
        # or genuinely quiet), emit one heartbeat row so log analysis sees the run.
        rows_to_emit = list(digests.items()) or [("(none)", Digest(recipient="(none)"))]
        for r, d in rows_to_emit:
            titles = " | ".join(f.reminder.title for f in d.fires)
            w.writerow([
                today.isoformat(), r,
                len(d.fires), len(d.dropped),
                skipped_tombstone, "yes" if dry_run else "no",
                titles,
            ])


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run(today: date, dry_run: bool = False, now: datetime | None = None,
        sheet_path: Path | None = None) -> dict[str, Digest]:
    """`now` is the wall-clock moment used for the tombstone window. For
    --as-of runs we anchor it at DAILY_RUN_TIME (07:30 local) on that date so
    fixtures are deterministic; real runs default to datetime.now()."""
    if now is None:
        now = datetime.now() if today == date.today() else datetime.combine(today, time(7, 30))
    path = sheet_path or SHEET_PATH

    reminders = read_reminders(path)

    # Phase 6.1 — tombstone guard runs BEFORE classify. Spec step 2a.
    active: list[Reminder] = []
    tombstoned: list[tuple[Reminder, float]] = []  # (reminder, age_hours)
    for r in reminders:
        if is_tombstoned(r, now):
            age_h = (now - r.write_queue_tombstone).total_seconds() / 3600.0
            tombstoned.append((r, age_h))
        else:
            active.append(r)

    fires = [f for f in (classify(r, today) for r in active) if f]
    digests = route(fires)
    for d in digests.values():
        apply_budget(d)

    if tombstoned:
        print(f"tombstone-guard: skipped {len(tombstoned)} row(s) "
              f"(window={TOMBSTONE_SKIP_HOURS}h)")
        for r, age_h in tombstoned:
            print(f"  row {r.row} '{r.title}' — tombstone age {age_h:.2f}h")

    if dry_run:
        print(f"--- DRY RUN for {today} ---")
        if not digests:
            print("(no recipients — quiet day)")
        for r, d in digests.items():
            print(f"\n=== to {r} ===")
            print(render_digest(d, today))
        append_log(today, digests, LOG_PATH, skipped_tombstone=len(tombstoned),
                   dry_run=True)
        return digests

    if not digests:
        # Heartbeat: still emit one quiet-day briefing so we know the engine ran
        quiet = Digest(recipient="Adar")
        out = write_briefing(quiet, today, BRIEFINGS_DIR)
        print(f"wrote {out} (quiet day)")
        digests = {"Adar": quiet}
    else:
        for d in digests.values():
            out = write_briefing(d, today, BRIEFINGS_DIR)
            print(f"wrote {out}")
    append_log(today, digests, LOG_PATH, skipped_tombstone=len(tombstoned))
    return digests


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--as-of", help="YYYY-MM-DD; defaults to today")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    today = datetime.strptime(args.as_of, "%Y-%m-%d").date() if args.as_of else date.today()
    run(today, dry_run=args.dry_run)


if __name__ == "__main__":
    main()
