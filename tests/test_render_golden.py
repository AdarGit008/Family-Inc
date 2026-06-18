"""Golden-file tests — daily digest + weekly briefing rendered byte-exact
against tests/fixtures/*.md (ENGINEERING §7).

The goldens freeze the M2 copy: DESIGN §6 Hebrew daily digest, D-014 footer
removal (messages end with content, not instructions), Hebrew קבוצות section.
The weekly briefing fallback is still the as-built English markdown. Copy
changes update goldens DELIBERATELY: review the diff, then regenerate with

    python3 tests/test_render_golden.py --regen
"""

import sys
from datetime import date, datetime, timedelta
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

from automation import daily_digest
from automation.lib.sheet import Reminder
from automation.reminders_engine import Digest, Fire
from automation.weekly_briefing import render_briefing

FIXTURES = Path(__file__).parent / "fixtures"
TODAY = date(2026, 6, 12)  # a Friday — exercises the Hebcal line


# ---------------------------------------------------------------------------
# Deterministic inputs
# ---------------------------------------------------------------------------
def _rem(title, due, **kw):
    base = dict(row=2, title=title, domain="Car", owner="Both", due=due,
                lead_times=[7, 1], recurrence="One-off", status="Pending",
                last_sent=None, channel="WhatsApp", notes="")
    base.update(kw)
    return Reminder(**base)


def digest_multi() -> str:
    d = Digest(recipient="adar")
    d.fires = [
        Fire(_rem("Car annual test", TODAY - timedelta(days=3)), "OVERDUE", -3),
        Fire(_rem("Kid dentist appointment", TODAY, domain="Health"), "FIRE TODAY", 0),
        Fire(_rem("Home insurance renewal", TODAY + timedelta(days=6), domain="Contracts"), "WEEK OUT", 6),
    ]
    d.dropped = [Fire(_rem("Passport renewal", TODAY + timedelta(days=30)), "MONTH OUT", 30)]
    return daily_digest.render_digest(d, TODAY)


def digest_single() -> str:
    d = Digest(recipient="shanee")
    d.fires = [Fire(_rem("Tipat Halav visit", TODAY + timedelta(days=1),
                         domain="Health", notes="Bring the pinkas"), "FIRE TODAY", 1)]
    return daily_digest.render_digest(d, TODAY)


def digest_quiet() -> str:
    return daily_digest.render_digest(Digest(recipient="adar"), TODAY)


def _assembled_inputs(tmp_path):
    """Sheet + WA digest + deferred + fake hebcal for the full assembly."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reminders"
    ws.append(["Title", "Domain", "Owner", "Due Date", "Lead Times", "Recurrence",
               "Status", "Last Sent", "Channel", "Notes", "Days Until", "Auto-flag",
               "LastDoneBy", "DoneAt", "WriteQueue_Tombstone", "Guide URL"])
    ws.append(["Car annual test", "Car", "Adar", date(2026, 6, 9), "7,1", "Yearly", "Pending"])
    ws.append(["Kid dentist appointment", "Health", "Adar", TODAY, "7,1", "One-off", "Pending"])
    sheet = tmp_path / "sheet.xlsx"
    wb.save(sheet)

    briefings = tmp_path / "Briefings"
    briefings.mkdir()
    (briefings / f"whatsapp_digest_{TODAY.isoformat()}.md").write_text(
        "קבוצות (24ש׳):\n"
        "גן — מחר יום פירות, להביא פרי חתוך (הגננת, 22:14)\n",
        encoding="utf-8")

    deferred = [{"id": "wa-9", "to": "adar", "body": "ועד: תיקון מעלית מחר",
                 "source": "whatsapp_summarizer", "deferred_on": "2026-06-11"}]

    def fake_shabbat(d):
        return {"candle_lighting": "2026-06-12T19:26:00+03:00",
                "havdalah": "2026-06-13T20:37:00+03:00", "parasha": "פרשת השבוע"}

    return sheet, briefings, deferred, fake_shabbat


def digest_assembled(tmp_path) -> str:
    sheet, briefings, deferred, fake_shabbat = _assembled_inputs(tmp_path)
    assembly = daily_digest.assemble(
        TODAY, now=datetime(2026, 6, 12, 7, 30), sheet_path=sheet,
        briefings_dir=briefings, shabbat_times=fake_shabbat, deferred=deferred)
    return assembly.messages["adar"]


def weekly() -> str:
    wb = Workbook()
    defs = {
        "Calendar-Events": [
            [date(2026, 6, 14), "10:00", "11:00", "Dentist", "Adar", "", "Clinic"],
            [date(2026, 6, 16), "", "", "Gan year-end party", "Both", "", "הגן"],
        ],
        "Reminders": [
            ["Car annual test", "Car", "Adar", date(2026, 6, 9), "7,1", "Yearly", "Pending"],
            ["Home insurance renewal", "Contracts", "Both", date(2026, 6, 17), "14,7", "Yearly", "Pending"],
        ],
        "Finance-Budget": [
            ["Groceries", 3000, 3400],
            ["Rent", 5000, 5000],
        ],
        "Goals": [
            ["House fund", "Both", "", date(2026, 7, 1), "Top up savings", 55, date(2026, 6, 1), "In progress"],
        ],
        "Finance-Accounts": [],
        "People": [],
    }
    for idx, (name, rows) in enumerate(defs.items()):
        ws = wb.active if idx == 0 else wb.create_sheet(name)
        if idx == 0:
            ws.title = name
        for r, row_data in enumerate(rows, start=2):
            for c, val in enumerate(row_data, start=1):
                ws.cell(r, c, val)
    return render_briefing(wb, TODAY)


# ---------------------------------------------------------------------------
# The tests
# ---------------------------------------------------------------------------
def _check(name: str, rendered: str):
    golden = (FIXTURES / name).read_text(encoding="utf-8")
    assert rendered == golden, (
        f"{name} drifted from its golden file. If the change is deliberate "
        f"(template work is M2), regenerate: python3 tests/test_render_golden.py --regen"
    )


def test_digest_multi_golden(tmp_runtime):
    _check("digest_multi.md", digest_multi())


def test_digest_single_golden(tmp_runtime):
    _check("digest_single.md", digest_single())


def test_digest_quiet_golden(tmp_runtime):
    _check("digest_quiet.md", digest_quiet())


def test_digest_assembled_golden(tmp_runtime, tmp_path):
    _check("digest_assembled.md", digest_assembled(tmp_path))


def test_weekly_briefing_golden(tmp_runtime):
    _check("weekly_briefing.md", weekly())


# ---------------------------------------------------------------------------
# Hebcal candle line — fires on erev-Shabbat (Fridays) AND erev-chag (B2)
# ---------------------------------------------------------------------------
def _boom(_d):
    raise AssertionError("wrong Hebcal path taken")


def test_hebcal_line_friday_uses_shabbat_times():
    line = daily_digest._hebcal_line(
        date(2026, 6, 12),  # a Friday
        lambda d: {"candle_lighting": "2026-06-12T19:26:00+03:00",
                   "havdalah": "2026-06-13T20:37:00+03:00"},
        chag_candles=_boom)
    assert "🕯 הדלקת נרות 19:26" in line and "צאת שבת 20:37" in line


def test_hebcal_line_erev_chag_uses_chag_candles():
    erev = date(2026, 6, 10)            # a Wednesday — exercises the non-Friday path
    assert erev.weekday() != 4
    line = daily_digest._hebcal_line(
        erev, _boom,
        chag_candles=lambda d: {"candle_lighting": "2026-06-10T19:10:00+03:00",
                                "havdalah": "2026-06-11T20:15:00+03:00"})
    assert "🕯 הדלקת נרות 19:10" in line and "צאת החג 20:15" in line


def test_hebcal_line_plain_weekday_renders_nothing():
    line = daily_digest._hebcal_line(
        date(2026, 6, 10), _boom, chag_candles=lambda d: None)
    assert line == ""


# ---------------------------------------------------------------------------
# Deliberate regeneration
# ---------------------------------------------------------------------------
if __name__ == "__main__" and "--regen" in sys.argv:
    import tempfile

    from automation.lib import config as _config

    FIXTURES.mkdir(exist_ok=True)
    with tempfile.TemporaryDirectory() as td:
        # Hermetic regen: a real logs/*.csv on this machine must not leak into
        # the goldens (reminders_log → run/tombstone counts; llm_costs → ₪ spend
        # in the §System self-report line, B6).
        _config.REMINDERS_LOG = Path(td) / "reminders_log.csv"
        _config.LLM_COSTS_LOG = Path(td) / "llm_costs.csv"
        outputs = {
            "digest_multi.md": digest_multi(),
            "digest_single.md": digest_single(),
            "digest_quiet.md": digest_quiet(),
            "digest_assembled.md": digest_assembled(Path(td)),
            "weekly_briefing.md": weekly(),
        }
    for name, body in outputs.items():
        (FIXTURES / name).write_text(body, encoding="utf-8")
        print(f"regenerated fixtures/{name} ({len(body)} chars)")
