"""Tests for automation/weekly_briefing.py — section builders with mock
workbook data (renamed from test_sunday_briefing.py in M1; D-011 moved the
briefing to Saturday and the rename followed)."""

from datetime import date, timedelta

from openpyxl import Workbook

from automation.lib.config import WEEK_AHEAD_DAYS
from automation.weekly_briefing import (
    render_briefing,
    section_goals,
    section_hygiene,
    section_money,
    section_reminders_week,
    section_week_ahead,
)


def _make_wb(sheet_defs: dict) -> Workbook:
    """Create an in-memory openpyxl Workbook with named sheets populated from
    list-of-lists data (row 1 = header placeholders so row 2+ is data)."""
    wb = Workbook()
    for idx, (sheet_name, rows) in enumerate(sheet_defs.items()):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        for r, row_data in enumerate(rows, start=2):
            for c, val in enumerate(row_data, start=1):
                ws.cell(r, c, val)
    return wb


# ---------------------------------------------------------------------------
# section_week_ahead
# ---------------------------------------------------------------------------
class TestSectionWeekAhead:
    def test_empty_when_no_events(self):
        wb = _make_wb({"Calendar-Events": []})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Quiet week" in result

    def test_shows_events_in_range(self):
        events = [
            [date(2026, 6, 12), "10:00", "11:00", "Dentist", "Adar", "", "Clinic"],
            [date(2026, 6, 14), "", "", "Family lunch", "Both", "", "Home"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Dentist" in result
        assert "Family lunch" in result
        assert "10:00–11:00" in result

    def test_excludes_events_before_today(self):
        events = [
            [date(2026, 6, 8), "09:00", "10:00", "Past event", "Adar"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Past event" not in result

    def test_excludes_events_after_window(self):
        events = [
            [date(2026, 6, 20), "09:00", "10:00", "Future event", "Adar"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Future event" not in result


# ---------------------------------------------------------------------------
# section_reminders_week
# ---------------------------------------------------------------------------
class TestSectionRemindersWeek:
    def test_empty_when_no_reminders(self):
        wb = _make_wb({"Reminders": []})
        today = date(2026, 6, 10)
        upcoming, overdue = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "No upcoming" in upcoming
        assert "No overdue" in overdue

    def test_upcoming_within_week(self):
        reminders = [
            ["Dentist checkup", "Health", "Adar", date(2026, 6, 14), "14,7,1", "One-off", "Pending"],
            ["Car test", "Auto", "Adar", date(2026, 6, 12), "7,1", "Annual", "Pending"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Dentist checkup" in upcoming
        assert "Car test" in upcoming

    def test_overdue_items(self):
        reminders = [
            ["Late bill", "Finance", "Adar", date(2026, 6, 1), "7,1", "Monthly", "Pending"],
            ["Future", "Kids", "Adar", date(2026, 6, 20), "7,1", "One-off", "Pending"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        _, overdue = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Late bill" in overdue
        assert "overdue" in overdue.lower()

    def test_skips_done_and_skipped(self):
        reminders = [
            ["Done item", "Kids", "Adar", date(2026, 6, 12), "7,1", "One-off", "Done"],
            ["Skipped item", "Kids", "Adar", date(2026, 6, 12), "7,1", "One-off", "Skipped"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Done item" not in upcoming
        assert "Skipped item" not in upcoming

    def test_skips_templated_rows(self):
        reminders = [
            ["[Template] thing", "Kids", "Adar", date(2026, 6, 12)],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Template" not in upcoming


# ---------------------------------------------------------------------------
# section_money
# ---------------------------------------------------------------------------
class TestSectionMoney:
    def test_basic_totals(self):
        rows = [
            ["Groceries", 3000, 2500],
            ["Rent", 5000, 5000],
            ["Utilities", 1000, 1200],
        ]
        wb = _make_wb({"Finance-Budget": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "₪8,700" in result  # 2500+5000+1200
        assert "₪9,000" in result  # 3000+5000+1000
        assert "Utilities" in result  # over-budget category
        assert "Groceries" not in result  # under budget, not in top 3

    def test_no_over_budget(self):
        rows = [
            ["Groceries", 3000, 2000],
            ["Rent", 5000, 5000],
        ]
        wb = _make_wb({"Finance-Budget": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "No categories over budget" in result

    def test_skips_total_row(self):
        rows = [
            ["Groceries", 3000, 2500],
            ["TOTAL", 10000, 9000],
        ]
        wb = _make_wb({"Finance-Budget": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "TOTAL" not in result

    def test_no_mom_block_without_last_month_column(self):
        # The 3-col stubs (no header row / no MoM column) must render no MoM —
        # this is what keeps the golden fixture byte-identical (M6.4).
        wb = _make_wb({"Finance-Budget": [["Groceries", 3000, 2500]]})
        result = section_money(wb, date(2026, 6, 10))
        assert "vs. last month" not in result

    def test_month_over_month_when_last_month_column_present(self):
        # A header row with "Last Month (ILS)" at col 10 (the live seed layout)
        # turns on per-category MoM. Groceries 2000→3000 (▲50%), Transport
        # 1000→500 (▼50%).
        wb = Workbook()
        ws = wb.active
        ws.title = "Finance-Budget"
        headers = ["Category", "Monthly Target (ILS)", "Actual (current month)",
                   "Variance", "% of Target", "YTD Actual", "Notes",
                   "As-of date", "", "Last Month (ILS)"]
        for c, h in enumerate(headers, start=1):
            ws.cell(1, c, h)
        for r, (cat, tgt, act, prev) in enumerate(
                [("Groceries", 4000, 3000, 2000), ("Transport", 1500, 500, 1000)],
                start=2):
            ws.cell(r, 1, cat); ws.cell(r, 2, tgt)
            ws.cell(r, 3, act); ws.cell(r, 10, prev)
        result = section_money(wb, date(2026, 6, 10))
        assert "vs. last month" in result
        assert "▲ 50% from ₪2,000" in result
        assert "▼ 50% from ₪1,000" in result


# ---------------------------------------------------------------------------
# section_goals
# ---------------------------------------------------------------------------
class TestSectionGoals:
    def test_empty_goals(self):
        wb = _make_wb({"Goals": []})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "No goals tracked" in result

    def test_goal_with_milestone_flag(self):
        rows = [
            ["Lose 5kg", "Shanee", "", date(2026, 6, 25), "Start walking 3x/week", 40, date(2026, 6, 5), "In progress"],
        ]
        wb = _make_wb({"Goals": rows})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "Lose 5kg" in result
        assert "milestone in 15d" in result

    def test_stale_goal_warning(self):
        rows = [
            ["Read books", "Adar", "", date(2026, 12, 31), "", 10, date(2026, 5, 1), "In progress"],
        ]
        wb = _make_wb({"Goals": rows})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "no update" in result


# ---------------------------------------------------------------------------
# section_hygiene (tmp_runtime: hygiene reads the schema-drift + engine-flag
# files — on the appliance those exist for real, tests must not see them)
# ---------------------------------------------------------------------------
class TestSectionHygiene:
    def test_all_clean(self, tmp_runtime):
        wb = _make_wb({
            "Reminders": [["OK", "", "", date(2026, 6, 15)]],
            "Finance-Accounts": [["Bank", "", "", "", "", "", date(2026, 6, 1)]],
            "People": [["Adar"]],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "All clean" in result

    def test_schema_drift_flag_surfaces(self, tmp_runtime):
        """ENGINEERING §8: the briefing is where humans hear about drift."""
        import json
        from automation.lib import config
        config.SCHEMA_DRIFT_FLAG.parent.mkdir(parents=True, exist_ok=True)
        config.SCHEMA_DRIFT_FLAG.write_text(json.dumps(
            {"at": "2026-06-10T07:25:00", "tab": "Reminders",
             "problems": ["col 5: expected 'Lead Times', found 'Lead Tymes'"]}),
            encoding="utf-8")
        wb = _make_wb({"Reminders": [], "Finance-Accounts": [], "People": []})
        result = section_hygiene(wb, date(2026, 6, 10))
        assert "schema drift" in result
        assert "Lead Tymes" in result

    def test_engine_flags_surface(self, tmp_runtime):
        from automation.lib import config
        config.ENGINE_FLAGS.parent.mkdir(parents=True, exist_ok=True)
        config.ENGINE_FLAGS.write_text(
            '{"at": "2026-06-10T07:25:00", "reason": "recurrence_clamped_to_month_end", '
            '"row": 7, "title": "Lease anniversary"}\n', encoding="utf-8")
        wb = _make_wb({"Reminders": [], "Finance-Accounts": [], "People": []})
        result = section_hygiene(wb, date(2026, 6, 10))
        assert "recurrence_clamped_to_month_end" in result
        assert "Lease anniversary" in result

    def test_missing_due_date(self, tmp_runtime):
        wb = _make_wb({
            "Reminders": [["Missing due", "", "", None]],
            "Finance-Accounts": [],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "missing a due date" in result.lower()

    def test_stale_account(self, tmp_runtime):
        wb = _make_wb({
            "Reminders": [],
            "Finance-Accounts": [["Old Bank", "", "", "", "", "", date(2026, 1, 1)]],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "not imported in" in result

    def test_placeholder_people(self, tmp_runtime):
        wb = _make_wb({
            "Reminders": [],
            "Finance-Accounts": [],
            "People": [["[replace me]"]],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "placeholder" in result


# ---------------------------------------------------------------------------
# render_briefing (integration smoke)
# ---------------------------------------------------------------------------
class TestRenderBriefing:
    def test_renders_all_sections(self, tmp_runtime):
        wb = _make_wb({
            "Calendar-Events": [[date(2026, 6, 12), "10:00", "11:00", "Dentist", "Adar"]],
            "Reminders": [["Upcoming", "Health", "Adar", date(2026, 6, 14), "7", "One-off", "Pending"]],
            "Finance-Budget": [["Groceries", 3000, 2500]],
            "Goals": [["Goal 1", "Adar", "", None, "", 0, None, "Not started"]],
            "Finance-Accounts": [],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = render_briefing(wb, today)
        assert "Family inc. — Weekly Briefing" in result
        assert "Week ahead" in result
        assert "Reminders firing this week" in result
        assert "Overdue" in result
        assert "Money" in result
        assert "Goals" in result
        assert "Data hygiene" in result
        assert f"week of {today.isoformat()}" in result
