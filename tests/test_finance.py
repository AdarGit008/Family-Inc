"""Tests for automation/finance_ingest.py (SPEC §12.2, M6.1; D-049/050/051).

Hermetic: a mock per-provider CSV → ingest → a tmp xlsx Sheet (explicit path,
never the live Sheet / committed seed). No banks, no Node, no network — the
Node scraper (scrape.js) is VPS-only and node-checked, not unit-tested.
Covers: CSV → Sheet, Txn-ID dedup + rerun idempotency, balance-only rows,
natural-key hash ids (the provider `identifier` is ignored — non-unique on
Mizrahi), Finance-Accounts upsert (human fields preserved),
fail-loud on missing creds / scrape-error marker, and the seed-safety gate.
"""

import shutil
import subprocess
from datetime import date, datetime
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from automation import finance_ingest as fin
from automation.lib import config as cfg
from automation.lib import sheet

SCRAPE_JS = Path(__file__).resolve().parents[1] / "automation" / "finance" / "scrape.js"

TODAY = date(2026, 6, 17)
NOW = datetime(2026, 6, 17, 6, 0, 0)

CSV = (
    "account,balance,date,identifier,amount,description\n"
    "MIZ-0001,12500.00,2026-06-15,abc123,-432.50,SHUFERSAL DEAL\n"
    "MIZ-0001,12500.00,2026-06-16,,-89.90,SUPERPHARM\n"
    "MIZ-0777,,2026-06-15,,-280.00,PAZ GAS\n"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _csv(tmp_path, text=CSV, name="mizrahi_2026-06-17.csv"):
    d = tmp_path / "stage"
    d.mkdir(parents=True, exist_ok=True)
    (d / name).write_text(text, encoding="utf-8")
    return d


def _sheet(tmp_path, accounts_rows=None):
    """A tmp xlsx with a placeholder tab (so the file loads); optionally a
    pre-seeded Finance-Accounts tab to test upsert-preserves-human-fields."""
    wb = Workbook()
    wb.active.title = "Placeholder"
    if accounts_rows is not None:
        ws = wb.create_sheet(cfg.FINANCE_ACCOUNTS_TAB)
        ws.append(sheet.FINANCE_ACCOUNTS_COLUMNS)
        for r in accounts_rows:
            ws.append(r)
    p = tmp_path / "finance.xlsx"
    wb.save(p)
    return p


def _rows(path, tab):
    wb = load_workbook(path, data_only=True)
    if tab not in wb.sheetnames:
        return []
    return [[c.value for c in row] for row in wb[tab].iter_rows()]


def _col(tab_rows, name):
    """Values under a named column (skips header)."""
    i = tab_rows[0].index(name)
    return [r[i] for r in tab_rows[1:]]


# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------
def test_provider_of():
    from pathlib import Path
    assert fin.provider_of(Path("mizrahi_2026-06-17.csv")) == "mizrahi"
    assert fin.provider_of(Path("/var/x/MAX_2026-06-17.csv")) == "max"


def test_txn_id_is_natural_key_hash():
    # The provider identifier is NOT part of the key (Mizrahi reuses it across
    # distinct charges — §12.2). Same natural key → same id; a different natural
    # key (here, a different amount) → a different id. Deterministic.
    h1 = fin.txn_id("2026-06-15", -10, "X", "A")
    h2 = fin.txn_id("2026-06-15", -10, "X", "A")
    h3 = fin.txn_id("2026-06-15", -11, "X", "A")
    assert h1.startswith("h:") and h1 == h2 and h1 != h3


# ---------------------------------------------------------------------------
# CSV → Sheet
# ---------------------------------------------------------------------------
def test_mock_csv_ingests_to_sheet(tmp_path):
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=_csv(tmp_path), sheet_path=sp, today=TODAY, now=NOW)
    assert res.txns_new == 3 and res.accounts == 2 and res.wrote

    txns = _rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)
    assert txns[0] == sheet.FINANCE_TRANSACTIONS_COLUMNS          # header
    assert len(txns) - 1 == 3
    # All Txn-IDs are natural-key hashes — the provider `identifier` (abc123 on
    # the first row) is ignored, since it is not unique per transaction (§12.2).
    assert all(str(t).startswith("h:") for t in _col(txns, "Txn-ID"))
    assert len(set(_col(txns, "Txn-ID"))) == 3          # 3 distinct natural keys
    # M6.4: the on-box rules engine categorizes at ingest (no LLM key in tests
    # → rules only). SHUFERSAL→Groceries, SUPERPHARM→Health, PAZ→Transport.
    cat_by_desc = dict(zip(_col(txns, "Description"), _col(txns, "Category")))
    assert cat_by_desc["SHUFERSAL DEAL"] == "Groceries"
    assert cat_by_desc["SUPERPHARM"] == "Health"
    assert cat_by_desc["PAZ GAS"] == "Transport"
    assert set(_col(txns, "Cat-Source")) == {"rules"}
    assert "abc123" not in _col(txns, "Txn-ID")         # identifier is not the key
    assert -432.5 in _col(txns, "Amount (ILS)")

    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    assert len(accts) - 1 == 2
    miz1 = [r for r in accts[1:] if r[0] == "MIZ-0001"][0]
    assert miz1[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Type")] == "bank"
    assert miz1[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Balance Snapshot")] == 12500.0
    assert miz1[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Last Imported")] == "2026-06-17"
    assert miz1[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Last 4")] == "0001"


def test_dedup_rerun_is_idempotent(tmp_path):
    sp = _sheet(tmp_path)
    d = _csv(tmp_path)
    fin.run(csv_dir=d, sheet_path=sp, today=TODAY, now=NOW)
    res2 = fin.run(csv_dir=d, sheet_path=sp, today=TODAY, now=NOW)
    assert res2.txns_new == 0 and res2.txns_seen == 3
    assert len(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)) - 1 == 3   # no dupes
    assert len(_rows(sp, cfg.FINANCE_ACCOUNTS_TAB)) - 1 == 2       # upsert, not append


def test_natural_key_collision_drops_second_charge_as_phantom_dup(tmp_path):
    """Accepted floor (finance_ingest.py txn_id docstring): two genuinely distinct
    same-day charges with an identical account+amount+description hash to one
    Txn-ID, so the dedup drops the second as a phantom dup. This LOCKS that
    behavior (it is silent data loss) — recovering both needs a richer key + a PO
    call. Rare for a bank; re-verify per card before trusting a provider field."""
    text = ("account,balance,date,identifier,amount,description\n"
            "MIZ-0001,9000,2026-06-15,,-12.00,COFFEE KIOSK\n"
            "MIZ-0001,9000,2026-06-15,,-12.00,COFFEE KIOSK\n")   # two real ₪12 coffees
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=_csv(tmp_path, text), sheet_path=sp, today=TODAY, now=NOW)
    # finance-ingest#3: the in-batch collision is a distinct counter, NOT
    # mislabeled "already on the tab" (txns_seen).
    assert res.txns_new == 1 and res.txns_phantom_dup == 1 and res.txns_seen == 0
    assert len(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)) - 1 == 1


def test_distinct_identifiers_do_not_rescue_same_natural_key(tmp_path):
    """Inverts the pre-2026-06-19 behavior: distinct provider identifiers no
    longer keep two same-natural-key rows apart — `identifier` is out of the key
    (§12.2), so the second still drops. We stopped trusting identifier because
    Mizrahi reuses it; this same-day-collision merge is the symmetric cost."""
    text = ("account,balance,date,identifier,amount,description\n"
            "MIZ-0001,9000,2026-06-15,txn-a,-12.00,COFFEE KIOSK\n"
            "MIZ-0001,9000,2026-06-15,txn-b,-12.00,COFFEE KIOSK\n")
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=_csv(tmp_path, text), sheet_path=sp, today=TODAY, now=NOW)
    assert res.txns_new == 1 and res.txns_phantom_dup == 1
    assert all(str(t).startswith("h:")
               for t in _col(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB), "Txn-ID"))


def test_reused_identifier_across_distinct_charges_keeps_all(tmp_path):
    """Regression for the 2026-06-19 live data-loss bug: israeli-bank-scrapers
    hands Mizrahi a NON-unique identifier (one id shared across many distinct
    charges). The old `if identifier: return identifier` collapsed 96 real rows
    to 26 on the live tab. With the natural-key Txn-ID, distinct charges that
    happen to share an identifier are all kept (would FAIL on the old code:
    txns_new == 1, phantom_dup == 2)."""
    text = ("account,balance,date,identifier,amount,description\n"
            "MIZ-0001,9000,2026-06-10,dup-ref,-12.00,COFFEE\n"
            "MIZ-0001,9000,2026-06-11,dup-ref,-50.00,GROCERY\n"
            "MIZ-0001,9000,2026-06-12,dup-ref,-9.90,BAKERY\n")   # 3 distinct, 1 id
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=_csv(tmp_path, text), sheet_path=sp, today=TODAY, now=NOW)
    assert res.txns_new == 3 and res.txns_phantom_dup == 0       # all kept, no false dup
    assert len(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)) - 1 == 3


def test_balance_only_row_feeds_account_not_txn(tmp_path):
    sp = _sheet(tmp_path)
    text = ("account,balance,date,identifier,amount,description\n"
            "MIZ-0001,9000,2026-06-15,,-50,COFFEE\n"
            "SAV-0002,40000,,,,\n")          # balance-only: no date/amount
    res = fin.run(csv_dir=_csv(tmp_path, text), sheet_path=sp, today=TODAY, now=NOW)
    assert res.txns_new == 1 and res.accounts == 2
    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    sav = [r for r in accts[1:] if r[0] == "SAV-0002"][0]
    assert sav[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Balance Snapshot")] == 40000
    assert "SAV-0002" not in _col(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB), "Account")


def test_upsert_preserves_human_fields(tmp_path):
    # Pre-seed MIZ-0001 with human-edited Owner/Notes + a stale balance.
    pre = ["MIZ-0001", "bank", "mizrahi", "0001", "Adar", "ILS",
           "2026-01-01", 999, "my main account"]
    sp = _sheet(tmp_path, accounts_rows=[pre])
    fin.run(csv_dir=_csv(tmp_path), sheet_path=sp, today=TODAY, now=NOW)
    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    miz1 = [r for r in accts[1:] if r[0] == "MIZ-0001"][0]
    C = sheet.FINANCE_ACCOUNTS_COLUMNS
    assert miz1[C.index("Owner")] == "Adar"             # human field preserved
    assert miz1[C.index("Notes")] == "my main account"  # preserved
    assert miz1[C.index("Balance Snapshot")] == 12500.0  # refreshed
    assert miz1[C.index("Last Imported")] == "2026-06-17"  # refreshed
    # MIZ-0777 is new → appended (so 2 account rows total, no duplicate MIZ-0001)
    assert len(accts) - 1 == 2


# ---------------------------------------------------------------------------
# Fail-loud + seed-safety
# ---------------------------------------------------------------------------
def test_fail_loud_when_no_csvs_on_live(tmp_path):
    empty = tmp_path / "stage"
    empty.mkdir()
    with pytest.raises(fin.FinanceError, match="no finance CSVs"):
        fin.run(csv_dir=empty, today=TODAY, now=NOW, live_override=True)


def test_mock_mode_when_no_csvs_and_not_live(tmp_path):
    empty = tmp_path / "stage"
    empty.mkdir()
    res = fin.run(csv_dir=empty, today=TODAY, now=NOW, live_override=False)
    assert res.is_mock and not res.wrote


def test_scrape_error_marker_fails_loud_after_persisting(tmp_path):
    sp = _sheet(tmp_path)
    d = _csv(tmp_path)
    (d / fin.SCRAPE_ERRORS_FILE).write_text(
        '{"errors": [{"provider": "max", "error": "OTP re-challenge"}]}',
        encoding="utf-8")
    with pytest.raises(fin.FinanceError, match="OTP re-challenge"):
        fin.run(csv_dir=d, sheet_path=sp, today=TODAY, now=NOW)
    # the good CSV's transactions were persisted BEFORE the raise (no data lost)
    assert len(_rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)) - 1 == 3


def test_not_live_no_path_writes_nothing(tmp_path):
    # csvs present, but no live backend and no explicit path → parse, write
    # nothing (never touches the committed seed — D-038/M2 invariant).
    res = fin.run(csv_dir=_csv(tmp_path), today=TODAY, now=NOW, live_override=False)
    assert res.txns_new == 3 and not res.wrote


def test_dry_run_writes_nothing(tmp_path):
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=_csv(tmp_path), sheet_path=sp, today=TODAY, now=NOW,
                  dry_run=True)
    assert not res.wrote
    assert _rows(sp, cfg.FINANCE_TRANSACTIONS_TAB) == []   # tab never created


def test_amount_with_commas_and_iso_date(tmp_path):
    sp = _sheet(tmp_path)
    text = ("account,balance,date,identifier,amount,description\n"
            'MIZ-0001,"1,234.50",2026-06-15,,"-1,200.00",RENT\n')
    fin.run(csv_dir=_csv(tmp_path, text), sheet_path=sp, today=TODAY, now=NOW)
    txns = _rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)
    assert -1200.0 in _col(txns, "Amount (ILS)")
    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    assert accts[1][sheet.FINANCE_ACCOUNTS_COLUMNS.index("Balance Snapshot")] == 1234.5


# ---------------------------------------------------------------------------
# lib/sheet.upsert_rows — direct unit (seed-safety gate)
# ---------------------------------------------------------------------------
def test_upsert_rows_skips_without_live_or_path(tmp_path, monkeypatch):
    # No path + not live → must not write anything (won't touch the seed).
    monkeypatch.setattr(sheet, "is_live", lambda: False)
    sheet.upsert_rows("Whatever", ["Account Name", "Balance Snapshot"],
                      [{"Account Name": "X", "Balance Snapshot": 1}],
                      key_column="Account Name")   # path=None → no-op, no crash


# ---------------------------------------------------------------------------
# Schema contract — column ORDER is load-bearing (review S3, D-052)
# ---------------------------------------------------------------------------
def test_transactions_column_order_is_load_bearing():
    # The seed's Finance-Budget actuals are SUMIFS over Date(A)/Amount(D)/
    # Category(E). A reorder here silently breaks the live budget formulas —
    # pin it so that can't happen without a failing test.
    cols = sheet.FINANCE_TRANSACTIONS_COLUMNS
    assert cols[0] == "Date"          # column A
    assert cols[3] == "Amount (ILS)"  # column D
    assert cols[4] == "Category"      # column E


def test_upsert_creates_accounts_tab_when_absent(tmp_path):
    # The new-tab branch of upsert_rows (no pre-seeded Finance-Accounts).
    sp = _sheet(tmp_path)   # only a Placeholder tab exists
    fin.run(csv_dir=_csv(tmp_path), sheet_path=sp, today=TODAY, now=NOW)
    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    assert accts[0] == sheet.FINANCE_ACCOUNTS_COLUMNS   # header created
    assert len(accts) - 1 == 2


def test_imported_at_is_populated_from_now(tmp_path):
    sp = _sheet(tmp_path)
    fin.run(csv_dir=_csv(tmp_path), sheet_path=sp, today=TODAY, now=NOW)
    txns = _rows(sp, cfg.FINANCE_TRANSACTIONS_TAB)
    assert all(v == NOW.isoformat(timespec="seconds")
               for v in _col(txns, "Imported-At"))


def test_multiple_provider_csvs_merge_in_one_run(tmp_path):
    d = _csv(tmp_path)   # mizrahi_… → MIZ-0001, MIZ-0777
    (d / "max_2026-06-17.csv").write_text(
        "account,balance,date,identifier,amount,description\n"
        "MAX-1234,,2026-06-15,maxid1,-150.00,SHOP\n", encoding="utf-8")
    sp = _sheet(tmp_path)
    res = fin.run(csv_dir=d, sheet_path=sp, today=TODAY, now=NOW)
    assert res.accounts == 3   # MIZ-0001, MIZ-0777, MAX-1234
    accts = _rows(sp, cfg.FINANCE_ACCOUNTS_TAB)
    max_row = [r for r in accts[1:] if r[0] == "MAX-1234"][0]
    assert max_row[sheet.FINANCE_ACCOUNTS_COLUMNS.index("Type")] == "card"


# ---------------------------------------------------------------------------
# Categorization — on-box rules + DeepSeek gap-fill (M6.4, §8.6; D-050/051)
# ---------------------------------------------------------------------------
from automation.lib import categorize  # noqa: E402
from automation.lib import llm  # noqa: E402


def test_rules_engine_maps_known_merchants():
    rules = categorize.load_rules()
    assert categorize.apply_rules("SHUFERSAL DEAL TLV", rules) == "Groceries"
    assert categorize.apply_rules("פז חיפה דרום", rules) == "Transport"
    # Ordering is load-bearing: SUPERPHARM must resolve to Health, not Groceries.
    assert categorize.apply_rules("SUPERPHARM 123", rules) == "Health"
    assert categorize.apply_rules("totally unknown vendor", rules) is None


def test_rules_vocabulary_is_distinct_and_seeded():
    vocab = categorize.vocabulary(categorize.load_rules())
    assert {"Groceries", "Health", "Transport"} <= set(vocab)
    assert len(vocab) == len(set(vocab))            # no dupes — the LLM vocab


def test_rules_vocab_within_budget_categories():
    """GAP-1 (single highest-value audit fix): every category the rules engine /
    LLM may emit MUST be a Finance-Budget row, or that category's actuals SUMIFS
    reads ₪0 and the spend is invisible. 'Dining'→'Dining out' is now aligned.
    KNOWN-PENDING — Shanee's budget-vocab migration is the authority: Fees/Income/
    Shopping have no budget row yet, held as an explicit allow-list so this test
    still catches any NEW drift and the gap can't be forgotten. When Shanee maps
    them (add rows, or remap the rules), shrink `pending` toward empty. The EXCLUDED
    set ('Card Settlement', `categorize.EXCLUDED_CATEGORIES`) is the opposite of
    pending: a label a RULE may assign (the Cal-settlement mirror) that must NEVER be
    a budget row AND must never be offered to the LLM gap-fill — so the spend counts
    once via the per-merchant Cal scrape, and the model can't zero a non-Cal row by
    guessing it."""
    rules = categorize.load_rules()
    all_cats = {cat for _, cat in rules}            # every label a RULE can assign (incl. excluded)
    llm_vocab = set(categorize.vocabulary(rules))   # the stage-2 LLM vocab (excludes the buckets)
    budget = {r[0] for r in _rows(cfg.SHEET_PATH, cfg.FINANCE_BUDGET_TAB)[1:]
              if r[0] and r[0] != "TOTAL"}
    pending = {"Fees", "Income", "Shopping"}        # GAP-1: awaiting Shanee's budget rows
    excluded = categorize.EXCLUDED_CATEGORIES       # 'Card Settlement' — rule-only, never a budget row
    unmapped = all_cats - budget - pending - excluded
    assert not unmapped, f"rules emit categories with no Finance-Budget row: {unmapped}"
    assert pending <= llm_vocab                     # the LLM may still emit the pending labels
    assert excluded <= all_cats                     # the exclusion bucket is defined as a rule
    assert excluded.isdisjoint(llm_vocab), (        # but NEVER in the LLM vocab (else it zeros a row)
        f"excluded buckets must not be offered to gap-fill: {excluded & llm_vocab}")
    assert not (excluded & budget), (               # and MUST stay out of the budget grid
        f"excluded categories must NOT be Finance-Budget rows: {excluded & budget}")
    assert "Dining out" in all_cats and "Dining" not in all_cats   # the GAP-1 rename


def test_card_settlement_excludes_cal_mirror():
    """Immediate-debit cards (Cal, and Shanee's Cal-cleared debit card) post each
    purchase per-merchant in the card's own scrape AND as a merchant-less settlement
    line on the Mizrahi debit. Those merchant-less mirror lines fall through to the
    EXCLUDED 'Card Settlement' bucket (not a budget row → out of the SUMIFS), so the
    spend counts once via the per-merchant card row. The exclusion block sits BELOW
    every merchant rule (a last-resort fallback), so a settlement token carrying a
    merchant suffix categorizes by its MERCHANT, never force-excluded — the M6.5
    box-verify (ii) over-match, closed structurally (2026-06-25). Tokens verified
    against live data: the כא"ל settlements + the future-charge line, and Shanee's
    'רכישה בכרטיס דביט' mirror (M6.5 2026-06-23, on the connected Cal login)."""
    rules = categorize.load_rules()
    # Genuinely merchant-less wrappers fall through to the exclusion bucket.
    assert categorize.apply_rules('דביט כא"ל (חיוב מיידי)', rules) == "Card Settlement"
    assert categorize.apply_rules('ויזה כא"ל (י)', rules) == "Card Settlement"
    assert categorize.apply_rules("חיוב ויזה כאל עתידי", rules) == "Card Settlement"
    # Shanee's debit card: its per-merchant detail rides the existing Cal scrape, so
    # the merchant-less Mizrahi mirror line is EXCLUDED — flipped from the 06-23 morning
    # guard (when her card wasn't yet scraped and the line was left in the budget).
    assert categorize.apply_rules("רכישה בכרטיס דביט", rules) == "Card Settlement"
    # CONTRACT (the fix for box-verify (ii)): a settlement token that carries a merchant
    # suffix categorizes by the MERCHANT, never the excluded bucket — because the block
    # sits below the merchant rules. A real grocery purchase is never silently zeroed.
    assert categorize.apply_rules("רכישה בכרטיס דביט שופרסל", rules) == "Groceries"
    assert categorize.apply_rules('ויזה כאל שופרסל', rules) == "Groceries"
    # Must NOT over-match: 'כארם' (Karem) has no merchant rule of its own → None, never
    # force-excluded; and a 'דמי כרטיס דביט' fee categorizes as Fees (the full 'רכישה ב…'
    # settlement phrase does not catch it).
    assert categorize.apply_rules("מסעדת ומאפיית כארם חסן", rules) != "Card Settlement"
    assert categorize.apply_rules("דמי כרטיס דביט", rules) == "Fees"
    # Deliberately absent from the budget grid (the SUMIFS exclusion) AND from the LLM
    # gap-fill vocab — reachable only by an exact settlement rule, never an LLM guess on
    # an ambiguous non-Cal row (which would silently zero a real expense).
    budget = {r[0] for r in _rows(cfg.SHEET_PATH, cfg.FINANCE_BUDGET_TAB)[1:] if r[0]}
    assert "Card Settlement" not in budget
    assert "Card Settlement" not in categorize.vocabulary(rules)


def test_excluded_bucket_never_shadows_a_merchant():
    """The Card Settlement exclusion is a LAST-RESORT fallback: it must sit below every
    merchant rule so a settlement wrapper that ever carries a merchant token categorizes
    by that MERCHANT, never the excluded bucket. Pins the file-order invariant — a future
    re-sort that lifts an excluded pattern above the merchant rules would silently zero
    real spend (the exact M6.5 over-match) — by appending each seeded merchant keyword to
    each excluded pattern and asserting the result is never the excluded bucket."""
    rules = categorize.load_rules()
    excluded = categorize.EXCLUDED_CATEGORIES
    excluded_pats = [pat for pat, cat in rules if cat in excluded]
    merchant_pats = [pat for pat, cat in rules if cat not in excluded]
    assert excluded_pats and merchant_pats            # guard: both populated
    for spat in excluded_pats:
        for mpat in merchant_pats:
            got = categorize.apply_rules(f"{spat} {mpat}", rules)
            assert got not in excluded, (
                f"'{spat} {mpat}' force-excluded as {got!r}: an excluded pattern is "
                f"shadowing merchant rule {mpat!r}. Keep the exclusion block BELOW all "
                f"merchant rules so a merchant-bearing line is never silently zeroed.")


def test_missing_rules_file_degrades_quiet(tmp_path):
    # No file → rules engine no-ops (returns []), categorize leaves blanks.
    txns = [{"Description": "SHUFERSAL", "Amount (ILS)": -10,
             "Category": "", "Cat-Source": ""}]
    categorize.categorize_transactions(
        txns, allow_llm=False, rules_path=tmp_path / "nope.csv")
    assert txns[0]["Category"] == "" and txns[0]["Cat-Source"] == ""


def test_unknown_stays_blank_without_llm(monkeypatch):
    monkeypatch.setattr(llm, "available", lambda: False)
    txns = [{"Description": "ZZZ MYSTERY", "Amount (ILS)": -10,
             "Category": "", "Cat-Source": ""}]
    categorize.categorize_transactions(txns, allow_llm=True)   # key-less → rules only
    assert txns[0]["Category"] == "" and txns[0]["Cat-Source"] == ""


def test_gapfill_sends_only_description_and_amount(monkeypatch):
    """§8.6 privacy: the gap-fill prompt carries description + amount only —
    never the account, the Txn-ID, or any other field of the row."""
    seen = {}

    def fake_complete(prompt, **kw):
        seen["prompt"] = prompt
        seen["system"] = kw.get("system", "")
        return '{"results":[{"i":0,"category":"Shopping"}]}'

    monkeypatch.setattr(llm, "available", lambda: True)
    monkeypatch.setattr(llm, "complete", fake_complete)
    txns = [{
        "Date": "2026-06-15", "Account": "MIZ-SECRET-9999",
        "Description": "MYSTERY VENDOR QX", "Amount (ILS)": -54.30,
        "Category": "", "Cat-Source": "",
        "Txn-ID": "secret-identifier-123", "Imported-At": "z",
    }]
    categorize.categorize_transactions(txns, allow_llm=True)
    assert txns[0]["Category"] == "Shopping" and txns[0]["Cat-Source"] == "llm"
    blob = seen["prompt"] + seen["system"]
    assert "MYSTERY VENDOR QX" in blob          # description: allowed
    assert "54.3" in blob                        # amount: allowed
    assert "MIZ-SECRET-9999" not in blob         # account: NEVER leaves the box
    assert "secret-identifier-123" not in blob   # Txn-ID: NEVER leaves the box


def test_gapfill_rejects_offvocab_answer(monkeypatch):
    # A category not in the rules vocab is dropped — the txn stays blank.
    monkeypatch.setattr(llm, "available", lambda: True)
    monkeypatch.setattr(llm, "complete",
                        lambda p, **k: '{"results":[{"i":0,"category":"Crypto"}]}')
    txns = [{"Description": "ZZZ MYSTERY", "Amount (ILS)": -10,
             "Category": "", "Cat-Source": ""}]
    categorize.categorize_transactions(txns, allow_llm=True)
    assert txns[0]["Category"] == "" and txns[0]["Cat-Source"] == ""


def test_gapfill_chunks_so_large_import_is_fully_categorized(monkeypatch):
    """B5: a rules-miss batch larger than GAPFILL_MAX_BATCH must be FULLY
    categorized before the write — chunk-looped, not truncated at the per-prompt
    cap. A blank-Category row keeps its real Txn-ID and is then excluded from
    dedup forever (never re-presented to the LLM), so an overflow left blank was
    permanent data loss on the first 45-day backlog."""
    import json
    budgets = []

    def fake_complete(prompt, **kw):
        budgets.append(kw.get("max_tokens"))
        # Cover every within-chunk index (a chunk is <= GAPFILL_MAX_BATCH rows).
        return json.dumps({"results": [{"i": i, "category": "Shopping"}
                                       for i in range(categorize.GAPFILL_MAX_BATCH)]})

    monkeypatch.setattr(llm, "available", lambda: True)
    monkeypatch.setattr(llm, "complete", fake_complete)
    n = categorize.GAPFILL_MAX_BATCH * 2 + 5    # 165 — well over one prompt's cap
    txns = [{"Description": f"ZZZ MYSTERY {i}", "Amount (ILS)": -10,
             "Category": "", "Cat-Source": ""} for i in range(n)]
    categorize.categorize_transactions(txns, allow_llm=True)
    assert all(t["Category"] == "Shopping" and t["Cat-Source"] == "llm" for t in txns)
    assert len(budgets) == 3                     # ceil(165/80) chunks, not one truncated call
    # A full 80-row chunk's reply (~1.5k tokens) must not be truncated by a fixed
    # small cap — the reply budget scales with the chunk (else the whole chunk
    # parses to {} and lands blank: the B5 data-loss in disguise).
    assert max(budgets) >= categorize.GAPFILL_MAX_BATCH * 16


# ---------------------------------------------------------------------------
# Budget reconciliation — the SUMIFS landmine (M6.4 build note; D-050)
# ---------------------------------------------------------------------------
def test_seed_budget_uses_text_prefix_not_serial_sumifs():
    """The landmine: a serial DATE() window over the RAW ISO-text Date column
    reads ₪0. Pin the seed's actuals to the locale-independent text-prefix form
    so the serial form can't silently return, and pin the MoM column."""
    b = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]
    c2 = b["C2"].value
    assert '$I$2&"*"' in c2               # current-month TEXT prefix
    assert '">="&DATE(' not in c2         # NOT the broken serial window
    assert 'TEXT($I$1,"yyyy")&"*"' in b["F2"].value          # YTD = year prefix
    assert b["J1"].value == "Last Month (ILS)"               # MoM column present
    assert b["I3"].value == '=TEXT(EDATE($I$1,-1),"yyyy-mm")'  # prev-month tag


def test_text_prefix_month_window_sums_iso_text_dates():
    """Prove the LOGIC the SUMIFS encodes works on the ISO-TEXT dates ingest
    writes. This Python re-implementation is DELIBERATE (tests-quality#3 / GAP-6):
    the seed backend reads formula cells as None offline (lib/sheet XlsxBackend
    data_only caveat), so the real SUMIFS can't be evaluated here — it's verified
    live at M6.3. This is the value that read ₪0 before the fix."""
    txns = [
        {"Date": "2026-06-03", "Category": "Groceries", "Amount (ILS)": -432.50},
        {"Date": "2026-06-20", "Category": "Groceries", "Amount (ILS)": -100.00},
        {"Date": "2026-05-28", "Category": "Groceries", "Amount (ILS)": -999.00},
        {"Date": "2026-06-10", "Category": "Transport", "Amount (ILS)": -280.00},
    ]

    def month_actual(cat, tag):   # mirrors -SUMIFS(D, E=cat, A like tag&"*")
        return -sum(t["Amount (ILS)"] for t in txns
                    if t["Category"] == cat and t["Date"].startswith(tag))

    assert month_actual("Groceries", "2026-06") == 532.50   # non-zero, this month
    assert month_actual("Groceries", "2026-05") == 999.00   # prev month isolated
    assert month_actual("Transport", "2026-06") == 280.00


# ---------------------------------------------------------------------------
# Budget formula INSTALLER — lib/finance_budget + finance_budget_formulas (M6.3,
# the M6.4 reconciliation step gated to live data). Single source of formula
# text; pinned against the committed seed so seed/installer/live can't diverge.
# ---------------------------------------------------------------------------
from automation.lib import finance_budget as fb  # noqa: E402
from automation import finance_budget_formulas as fbf  # noqa: E402

BUDGET_HEADER = ["Category", "Monthly Target (ILS)", "Actual (current month)",
                 "Variance", "% of Target", "YTD Actual", "Notes", "As-of date",
                 None, "Last Month (ILS)"]   # col I (9) header is the =TODAY() helper


def _budget_grid(categories=("Groceries", "Transport"), total=True, header=None):
    """A Finance-Budget grid (list[list]) — header + category rows (A name, B
    target) + an optional TOTAL row, machine columns blank (pre-install)."""
    g = [list(header if header is not None else BUDGET_HEADER)]
    for name in categories:
        row = [None] * 10
        row[fb.COL_CATEGORY - 1], row[fb.COL_TARGET - 1] = name, 1000
        g.append(row)
    if total:
        row = [None] * 10
        row[fb.COL_CATEGORY - 1] = "TOTAL"
        g.append(row)
    return g


def _budget_sheet(tmp_path, categories=(("Groceries", 4000), ("Transport", 1500)),
                  total=True, header=None):
    """A tmp xlsx with a Finance-Budget tab (header + categories + optional TOTAL),
    machine columns blank — the live tab's state before the installer runs."""
    wb = Workbook()
    wb.active.title = "Placeholder"
    ws = wb.create_sheet(cfg.FINANCE_BUDGET_TAB)
    ws.append(header if header is not None else BUDGET_HEADER)
    for name, target in categories:
        ws.append([name, target])
    if total:
        ws.append(["TOTAL"])
    p = tmp_path / "budget.xlsx"
    wb.save(p)
    return p


def test_budget_cells_are_text_prefix_not_serial():
    """The landmine guard at the installer level: the month/YTD/last-month actuals
    are TEXT-prefix wildcards over the ISO-text Date, never a serial DATE() window."""
    cells = {(r, c): v for r, c, v in fb.budget_formula_cells(_budget_grid())}
    c2 = cells[(2, fb.COL_ACTUAL)]
    assert '$I$2&"*"' in c2 and '">="&DATE(' not in c2          # month text-prefix
    assert 'TEXT($I$1,"yyyy")&"*"' in cells[(2, fb.COL_YTD)]    # YTD = year prefix
    assert '$I$3&"*"' in cells[(2, fb.COL_LASTMONTH)]           # last-month
    assert cells[(1, fb.COL_HELPER)] == "=TODAY()"
    assert cells[(2, fb.COL_HELPER)] == '=TEXT(I1,"yyyy-mm")'
    assert cells[(3, fb.COL_HELPER)] == '=TEXT(EDATE($I$1,-1),"yyyy-mm")'
    assert cells[(2, fb.COL_VARIANCE)] == "=B2-C2"
    assert cells[(2, fb.COL_PCT)] == "=IFERROR(C2/B2,0)"
    # TOTAL sums over the category span (rows 2..3), plus its own variance/%.
    assert cells[(4, fb.COL_ACTUAL)] == "=SUM(C2:C3)"
    assert cells[(4, fb.COL_TARGET)] == "=SUM(B2:B3)"
    assert cells[(4, fb.COL_YTD)] == "=SUM(F2:F3)"


def test_budget_cells_match_committed_seed():
    """The anti-drift tie: the installer's output for the seed's own categories
    must EQUAL the seed's pinned formulas, so a live install and the committed
    seed (test_seed_budget_uses_text_prefix_not_serial_sumifs) stay identical."""
    seed = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]   # formulas (not data_only)
    grid = [[seed.cell(row=r, column=c).value for c in range(1, 11)]
            for r in range(1, seed.max_row + 1)]
    cells = {(r, c): v for r, c, v in fb.budget_formula_cells(grid)}
    for coord, key in {"C2": (2, 3), "F2": (2, 6), "J2": (2, 10), "D2": (2, 4),
                       "E2": (2, 5), "I1": (1, 9), "I2": (2, 9), "I3": (3, 9),
                       "C13": (13, 3), "B13": (13, 2), "F13": (13, 6)}.items():
        assert cells[key] == seed[coord].value, coord
    # …and the WHOLE installer output equals the seed, not just the spot-checks —
    # any divergence at any of the 66 machine cells (e.g. YTD on rows 3-12, the
    # TOTAL variance/%, the H labels) fails here, making the anti-drift tie total.
    for r, c, v in fb.budget_formula_cells(grid):
        assert v == seed.cell(row=r, column=c).value, (r, c)


def test_budget_installer_round_trips_to_live_formulas(tmp_path):
    """The write seam: cells stamped via sheet.write_cells (USER_ENTERED) land as
    live formulas, and the human columns (Category/Target/Notes) are untouched."""
    sp = _budget_sheet(tmp_path)
    cells = fb.budget_formula_cells(sheet.read_grid(cfg.FINANCE_BUDGET_TAB, sp))
    sheet.write_cells(cfg.FINANCE_BUDGET_TAB, cells, path=sp)
    ws = load_workbook(sp)[cfg.FINANCE_BUDGET_TAB]                 # data_only=False → formulas
    assert ws["C2"].value.startswith("=IFERROR(-SUMIFS(") and '$I$2&"*"' in ws["C2"].value
    assert 'TEXT($I$1,"yyyy")&"*"' in ws["F2"].value
    assert '$I$3&"*"' in ws["J2"].value
    assert ws["I1"].value == "=TODAY()" and ws["I2"].value == '=TEXT(I1,"yyyy-mm")'
    assert ws["A2"].value == "Groceries" and ws["B2"].value == 4000   # human cols intact
    assert ws["G2"].value in (None, "")                               # Notes never written
    assert ws["C4"].value == "=SUM(C2:C3)"                            # TOTAL over 2 rows


def test_budget_installer_never_writes_human_columns():
    """The irreversible-harm guard: the installer must never write a category row's
    Category (A) or Monthly Target (B), or any Notes (G) — those are Shanee's. (The
    TOTAL row's B is a machine =SUM and IS allowed.) This pins all three prongs of
    the safety contract, so a future refactor that emitted a human cell fails here
    rather than clobbering the live budget on the next re-stamp. Also covers the old
    'stray Transport-row Notes SUMIFS' artifact class (the G prong)."""
    grid = _budget_grid(categories=("Housing", "Groceries", "Transport"))
    cats = set(fb.category_rows(grid))
    cells = fb.budget_formula_cells(grid)
    assert all(c != fb.COL_NOTES for _, c, _ in cells)                        # G never (any row)
    assert all(c != fb.COL_CATEGORY for _, c, _ in cells)                     # A never (any row)
    assert all(not (c == fb.COL_TARGET and r in cats) for r, c, _ in cells)   # B never on a category row
    assert any(c == fb.COL_TARGET for _, c, _ in cells)                       # …but the TOTAL B SUM is present


def test_budget_header_drift_fails_loud(tmp_path):
    """A renamed load-bearing column → refuse to stamp by position (fail loud,
    never guess which column the actuals belong in). A *machine* header holding a
    DIFFERENT value is a real column shift, distinct from an absent one (titled)."""
    bad = list(BUDGET_HEADER)
    bad[fb.COL_ACTUAL - 1] = "Spent"                  # C header drifted (non-empty conflict)
    with pytest.raises(fb.BudgetHeaderError):
        fb.budget_formula_cells(_budget_grid(header=bad))


def test_budget_installer_titles_absent_machine_header():
    """Hardening (2026-06-20): a tab that simply LACKS a machine column header —
    the live tab created before the M6.4 block had no J 'Last Month (ILS)', and a
    fresh budget from Shanee's migration has only Category/Target — must NOT fail
    loud. The installer owns those columns, so it titles them (row 1) and stamps the
    data. An absent header is no contradiction; a CONFLICTING one still refuses
    (test_budget_header_drift_fails_loud). Removes the manual 'add J1 by hand' step."""
    hdr = list(BUDGET_HEADER)
    hdr[fb.COL_LASTMONTH - 1] = None    # J absent — the exact live 2026-06-20 case
    hdr[fb.COL_YTD - 1] = ""            # F absent too (blank, not None)
    cells = {(r, c): v for r, c, v in fb.budget_formula_cells(_budget_grid(header=hdr))}
    assert cells[(1, fb.COL_LASTMONTH)] == "Last Month (ILS)"   # installer titled J
    assert cells[(1, fb.COL_YTD)] == "YTD Actual"               # …and F
    assert '$I$3&"*"' in cells[(2, fb.COL_LASTMONTH)]           # …and still stamped its data
    assert (1, fb.COL_CATEGORY) not in cells                    # never (re)titles a human header


def test_budget_absent_human_header_still_fails_loud():
    """The installer titles its OWN columns, never the human ones: an absent
    Category or Monthly Target header is a malformed tab it refuses, rather than
    silently re-titling Shanee's columns (the human-vs-machine ownership boundary)."""
    for human_col in (fb.COL_CATEGORY, fb.COL_TARGET):
        hdr = list(BUDGET_HEADER)
        hdr[human_col - 1] = None
        with pytest.raises(fb.BudgetHeaderError):
            fb.budget_formula_cells(_budget_grid(header=hdr))


def test_budget_no_categories_fails_loud():
    """A header-only tab (no budget rows yet) fails loud rather than stamping an
    empty layout — Shanee's migration must populate column A first."""
    with pytest.raises(fb.BudgetHeaderError):
        fb.budget_formula_cells(_budget_grid(categories=(), total=False))


def test_budget_category_below_total_fails_loud():
    """A category row ordered BELOW the TOTAL row would put TOTAL inside its own SUM
    range (circular #ERROR, and two surfaces read TOTAL) — refuse the layout, never
    emit a self-referential sum. Guards a re-run after a budget reorder."""
    grid = [list(BUDGET_HEADER), [None] * 10, [None] * 10, [None] * 10]
    grid[1][fb.COL_CATEGORY - 1], grid[1][fb.COL_TARGET - 1] = "Housing", 8500   # row 2
    grid[2][fb.COL_CATEGORY - 1] = "TOTAL"                                       # row 3
    grid[3][fb.COL_CATEGORY - 1], grid[3][fb.COL_TARGET - 1] = "Savings", 1000   # row 4 < TOTAL
    with pytest.raises(fb.BudgetHeaderError):
        fb.budget_formula_cells(grid)


def test_write_cells_skips_without_live_or_path(monkeypatch):
    """sheet.write_cells refuses to write when path is None and not live — the
    lib-level 'never mutate the committed seed' backstop (mirrors upsert_rows). The
    CLI short-circuits before reaching this branch, so pin it directly."""
    monkeypatch.setattr(sheet, "is_live", lambda: False)
    before = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]["C2"].value
    sheet.write_cells(cfg.FINANCE_BUDGET_TAB, [(2, 3, "=1+1")], path=None)   # no-op
    after = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]["C2"].value
    assert after == before                                                  # seed untouched


def test_installer_dry_run_writes_nothing(tmp_path, capsys):
    sp = _budget_sheet(tmp_path)
    cells = fbf.run(path=sp, dry_run=True)
    assert cells
    assert load_workbook(sp)[cfg.FINANCE_BUDGET_TAB]["C2"].value in (None, "")
    assert "dry-run" in capsys.readouterr().out


def test_installer_refuses_without_live_or_path(capsys):
    """No live backend + no path → builds the cells (reading the seed) but writes
    NOTHING, so a creds-less dev run can't mutate the committed seed."""
    before = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]["C2"].value
    cells = fbf.run(path=None, dry_run=False)          # is_live() is False (conftest)
    assert cells and "NOT written" in capsys.readouterr().out
    after = load_workbook(cfg.SHEET_PATH)[cfg.FINANCE_BUDGET_TAB]["C2"].value
    assert after == before                             # seed untouched


# ---------------------------------------------------------------------------
# scrape.js — the Node scraper is VPS-only (banks + bundled Chromium), so these
# guard only the parts that need neither: it parses, and its argv/fail-loud
# contract holds. The deps (israeli-bank-scrapers, puppeteer) are required
# lazily, so these run green without `npm ci` having been done.
# ---------------------------------------------------------------------------
requires_node = pytest.mark.skipif(
    shutil.which("node") is None, reason="node not installed (scrape.js is VPS-only)"
)


def _node(*args, env=None):
    return subprocess.run(
        ["node", str(SCRAPE_JS), *args],
        capture_output=True, text=True, env=env, timeout=30,
    )


@requires_node
def test_scrape_js_node_check_parses():
    """The syntax guard the module docstring promises (no test existed before)."""
    res = subprocess.run(
        ["node", "--check", str(SCRAPE_JS)], capture_output=True, text=True, timeout=30
    )
    assert res.returncode == 0, res.stderr


@requires_node
@pytest.mark.parametrize("args", [["--auth"], ["--auth", "bogus"]])
def test_scrape_js_auth_usage_guard(args):
    """`--auth` without a known provider exits 2 with usage and loads no Chromium
    (the guard runs before puppeteer is required, so it passes with no node_modules)."""
    res = _node(*args)
    assert res.returncode == 2
    assert "usage: node scrape.js --auth <provider>" in res.stderr
    assert "mizrahi, max, cal" in res.stderr


@requires_node
def test_scrape_js_auth_mizrahi_is_noop():
    """Mizrahi is password-only — `--auth mizrahi` is a clean no-op (exit 0), not a
    browser launch; only Max/Cal persist a device-trust profile."""
    res = _node("--auth", "mizrahi")
    assert res.returncode == 0
    assert "password-only" in res.stdout


@requires_node
@pytest.mark.parametrize("args", [["--foo"], ["mizrahi"], ["--auth", "max", "extra"]])
def test_scrape_js_rejects_unknown_invocation(args):
    """A typo or stray positional must fail loud (exit 2), never silently take the
    daily-scrape branch — fail-loud on the hand-typed `--auth` command."""
    res = _node(*args)
    assert res.returncode == 2
    assert "usage: node scrape.js" in res.stderr


@requires_node
def test_scrape_js_missing_creds_fails_loud():
    """A daily run with no creds file fails loud (exit 1) — the unit-fails →
    fail-flag contract. Lazy require means this hits loadCreds() before the lib."""
    import os
    env = {**os.environ, "FAMILY_INC_BANK_CREDS": "/nonexistent/bank_creds.json"}
    res = _node(env=env)
    assert res.returncode == 1
    assert "[fatal]" in res.stderr and "bank_creds.json" in res.stderr


# ---------------------------------------------------------------------------
# Re-categorize backfill + coverage (M6.4/M6.5 acceptance; SPEC §12.2)
# ---------------------------------------------------------------------------
from automation import finance_recategorize as recat  # noqa: E402
from automation.lib import finance_coverage as fincov  # noqa: E402

# Finance-Transactions rows for the backfill: a Cal mirror (→ Card Settlement via
# the existing rule), a grocery (→ Groceries), a genuine unknown (stays blank,
# LLM off in tests), Shanee's debit mirror (→ Card Settlement), and an
# already-categorized manual row whose description WOULD match a rule (PAZ →
# Transport) — it must survive untouched.
TXN_ROWS = [
    {"Date": "2026-06-15", "Account": "MIZ-0001", "Description": "ויזה כאל",
     "Amount (ILS)": -540.00, "Category": "", "Cat-Source": "",
     "Txn-ID": "h:cal0000000000a1", "Imported-At": "2026-06-19T06:00:00"},
    {"Date": "2026-06-15", "Account": "MIZ-0001", "Description": "SHUFERSAL DEAL TLV",
     "Amount (ILS)": -432.50, "Category": "", "Cat-Source": "",
     "Txn-ID": "h:groc000000000b2", "Imported-At": "2026-06-19T06:00:00"},
    {"Date": "2026-06-16", "Account": "MIZ-0001", "Description": "ZZZ MYSTERY VENDOR",
     "Amount (ILS)": -12.00, "Category": "", "Cat-Source": "",
     "Txn-ID": "h:unkn000000000c3", "Imported-At": "2026-06-19T06:00:00"},
    {"Date": "2026-06-16", "Account": "MIZ-0001", "Description": "רכישה בכרטיס דביט",
     "Amount (ILS)": -77.00, "Category": "", "Cat-Source": "",
     "Txn-ID": "h:dbit000000000d4", "Imported-At": "2026-06-19T06:00:00"},
    {"Date": "2026-06-17", "Account": "MIZ-0001", "Description": "PAZ GAS HAIFA",
     "Amount (ILS)": -280.00, "Category": "Health", "Cat-Source": "manual",
     "Txn-ID": "h:hand000000000e5", "Imported-At": "2026-06-19T06:00:00"},
]


def _txn_sheet(tmp_path, rows=TXN_ROWS, columns=None):
    """A tmp xlsx with a Finance-Transactions tab (header + rows)."""
    cols = columns or sheet.FINANCE_TRANSACTIONS_COLUMNS
    wb = Workbook()
    wb.active.title = "Placeholder"
    ws = wb.create_sheet(cfg.FINANCE_TRANSACTIONS_TAB)
    ws.append(cols)
    for r in rows:
        ws.append([r.get(c, "") for c in cols])
    p = tmp_path / "finance.xlsx"
    wb.save(p)
    return p


def _by_txn(path):
    """{Txn-ID: (Category, Cat-Source)} read back from the live tab. openpyxl
    rounds an empty cell to None on read — coerce to "" so a blank reads ("","")."""
    tab = _rows(path, cfg.FINANCE_TRANSACTIONS_TAB)
    hdr = tab[0]
    ti, ci, si = hdr.index("Txn-ID"), hdr.index("Category"), hdr.index("Cat-Source")
    s = lambda v: "" if v is None else v
    return {r[ti]: (s(r[ci]), s(r[si])) for r in tab[1:]}


def test_recategorize_backfills_blank_rows(tmp_path):
    sp = _txn_sheet(tmp_path)
    res = recat.run(sheet_path=sp, allow_llm=False)        # rules-only, deterministic
    assert res.wrote
    assert res.total == 5 and res.blank_before == 4
    assert res.recategorized == 3 and res.now_rules == 3 and res.still_blank == 1
    got = _by_txn(sp)
    assert got["h:cal0000000000a1"] == ("Card Settlement", "rules")  # Cal mirror
    assert got["h:groc000000000b2"] == ("Groceries", "rules")
    assert got["h:dbit000000000d4"] == ("Card Settlement", "rules")  # Shanee debit mirror
    assert got["h:unkn000000000c3"] == ("", "")                       # genuine unknown stays blank
    assert got["h:hand000000000e5"] == ("Health", "manual")           # manual row untouched


def test_recategorize_only_touches_blank_rows(tmp_path):
    """The manual 'Health' row whose description (PAZ) WOULD map to Transport is
    never re-derived — the backfill is scoped to blank rows, so a human (or prior)
    categorization is never clobbered."""
    sp = _txn_sheet(tmp_path)
    recat.run(sheet_path=sp, allow_llm=False)
    assert _by_txn(sp)["h:hand000000000e5"] == ("Health", "manual")


def test_recategorize_is_idempotent(tmp_path):
    sp = _txn_sheet(tmp_path)
    recat.run(sheet_path=sp, allow_llm=False)
    before = _by_txn(sp)
    res2 = recat.run(sheet_path=sp, allow_llm=False)       # only the 1 unknown left blank
    assert res2.blank_before == 1 and res2.recategorized == 0 and not res2.wrote
    assert _by_txn(sp) == before                            # nothing changed on the second pass


def test_recategorize_dry_run_writes_nothing(tmp_path):
    sp = _txn_sheet(tmp_path)
    res = recat.run(sheet_path=sp, dry_run=True, allow_llm=False)
    assert not res.wrote and res.recategorized == 3         # rules-preview counts the 3 hits
    # ...but the tab is untouched: every blank row is still blank.
    got = _by_txn(sp)
    for tid in ("h:cal0000000000a1", "h:groc000000000b2", "h:dbit000000000d4"):
        assert got[tid] == ("", "")


def test_recategorize_skips_without_live_or_path(monkeypatch):
    """Seed-safety: no live backend and no --sheet → reads/writes nothing, never
    touches the committed seed (mirrors test_upsert_rows_skips_without_live_or_path)."""
    monkeypatch.delenv(cfg.SHEET_ID_ENV, raising=False)
    sheet.reset_backend()
    res = recat.run(sheet_path=None)
    assert res.total == 0 and not res.wrote


def test_recategorize_fails_loud_on_missing_header(tmp_path):
    """A Finance-Transactions tab missing a load-bearing column (Cat-Source) fails
    loud (§7.1) rather than writing Category by guessed position."""
    cols = [c for c in sheet.FINANCE_TRANSACTIONS_COLUMNS if c != "Cat-Source"]
    sp = _txn_sheet(tmp_path, columns=cols)
    with pytest.raises(recat.RecategorizeError):
        recat.run(sheet_path=sp, allow_llm=False)


def test_recategorize_empty_tab_is_noop(tmp_path):
    sp = _txn_sheet(tmp_path, rows=[])
    res = recat.run(sheet_path=sp, allow_llm=False)
    assert res.total == 0 and not res.wrote


def test_recategorize_handles_blank_interior_row(tmp_path):
    """A fully-blank interior row must not shift the physical write index: every
    Txn-ID still maps to its OWN (Category, Cat-Source) and res.total counts only the
    non-blank rows. Pins the enumerate(grid[1:], start=2) invariant — a refactor that
    derived the write row from the filtered candidate list would stamp the rows BELOW
    the blank one onto the wrong transactions (silent ledger corruption)."""
    rows = TXN_ROWS[:2] + [{c: "" for c in sheet.FINANCE_TRANSACTIONS_COLUMNS}] + TXN_ROWS[2:]
    sp = _txn_sheet(tmp_path, rows=rows)
    res = recat.run(sheet_path=sp, allow_llm=False)
    assert res.total == 5                                  # the all-blank interior row is skipped
    got = _by_txn(sp)
    assert got["h:cal0000000000a1"] == ("Card Settlement", "rules")   # above the blank
    assert got["h:groc000000000b2"] == ("Groceries", "rules")
    assert got["h:dbit000000000d4"] == ("Card Settlement", "rules")   # below the blank — index held
    assert got["h:unkn000000000c3"] == ("", "")
    assert got["h:hand000000000e5"] == ("Health", "manual")           # below the blank — untouched


def test_recategorize_llm_fills_blank_and_dry_run_skips_llm(tmp_path, monkeypatch):
    """The live LLM gap-fill path: a genuine-unknown blank gets Cat-Source 'llm' from
    DeepSeek; and --dry-run must NOT call the LLM (the documented no-API-spend preview,
    so live merchant descriptions don't leave the box during a no-write run, §8.6)."""
    calls = {"n": 0}

    def fake_complete(prompt, **kw):
        calls["n"] += 1
        return '{"results":[{"i":0,"category":"Shopping"}]}'

    monkeypatch.setattr(llm, "available", lambda: True)
    monkeypatch.setattr(llm, "complete", fake_complete)
    rows = [r for r in TXN_ROWS if r["Txn-ID"] == "h:unkn000000000c3"]   # the one rules-miss
    sp = _txn_sheet(tmp_path, rows=rows)
    # (1) dry-run with allow_llm=True must NOT call the LLM and must not write.
    res_dry = recat.run(sheet_path=sp, dry_run=True, allow_llm=True)
    assert calls["n"] == 0 and not res_dry.wrote
    assert _by_txn(sp)["h:unkn000000000c3"] == ("", "")
    # (2) live run calls the LLM once and writes Cat-Source 'llm'.
    res = recat.run(sheet_path=sp, dry_run=False, allow_llm=True)
    assert calls["n"] == 1 and res.now_llm == 1 and res.wrote
    assert _by_txn(sp)["h:unkn000000000c3"] == ("Shopping", "llm")


# ---- coverage (the read-only yield surface) ----

COV_ROWS = [
    {"Category": "Groceries", "Cat-Source": "rules", "Account": "MIZ", "Description": "shufersal"},
    {"Category": "Health", "Cat-Source": "llm", "Account": "MIZ", "Description": "clinic"},
    {"Category": "Card Settlement", "Cat-Source": "rules", "Account": "MIZ", "Description": "ויזה כאל"},
    {"Category": "", "Cat-Source": "", "Account": "MIZ", "Description": "ATM WITHDRAWAL"},
    {"Category": "Dining out", "Cat-Source": "", "Account": "CAL", "Description": "wolt"},  # manual
]


def test_coverage_counts_and_buckets():
    c = fincov.coverage(COV_ROWS)
    assert c.total == 5
    assert c.categorized == 3 and c.excluded == 1 and c.blank == 1     # partition → total
    assert c.categorized + c.excluded + c.blank == c.total
    assert c.eligible == 4                                             # total − excluded
    assert round(c.coverage_pct, 3) == 0.75                            # 3/4
    # by_source is scoped to CATEGORIZED rows, so it sums to `categorized` and the
    # excluded Card-Settlement row (Cat-Source "rules") is NOT counted here — else the
    # rules sub-count would overshoot the categorized headline on live data.
    assert c.by_source == {"rules": 1, "llm": 1, "manual": 1}          # Groceries · Health · Dining out
    assert sum(c.by_source.values()) == c.categorized


def test_coverage_per_account():
    c = fincov.coverage(COV_ROWS)
    miz, cal = c.by_account["MIZ"], c.by_account["CAL"]
    assert (miz.total, miz.categorized, miz.excluded, miz.blank) == (4, 2, 1, 1)
    assert miz.eligible == 3 and round(miz.pct, 3) == round(2 / 3, 3)
    assert (cal.total, cal.categorized, cal.pct) == (1, 1, 1.0)


def test_coverage_blank_samples_name_the_merchant():
    c = fincov.coverage(COV_ROWS)
    assert c.blank_samples == [("ATM WITHDRAWAL", 1)]


def test_coverage_empty_degrades():
    c = fincov.coverage([])
    assert c.total == 0 and c.coverage_pct == 0.0 and c.eligible == 0
    assert "No finance transactions" in fincov.render(c, date(2026, 6, 26))


def test_coverage_rows_from_grid_skips_blank_and_short_rows():
    grid = [
        ["Date", "Account", "Description", "Category", "Cat-Source"],
        ["2026-06-15", "MIZ", "shufersal", "Groceries", "rules"],
        ["2026-06-16", "MIZ", "atm"],                 # short row — trailing cells absent
        [None, None, None, None, None],               # fully blank — skipped
    ]
    rows = fincov.rows_from_grid(grid)
    assert len(rows) == 2
    assert rows[0]["Category"] == "Groceries"
    assert rows[1]["Description"] == "atm" and rows[1]["Category"] is None


def test_coverage_cli_reads_live_tab(tmp_path):
    """The read-only CLI reads the tab through lib/sheet and never writes it."""
    from automation import finance_coverage as cli
    sp = _txn_sheet(tmp_path)                          # 5 rows, 4 blank, 0 categorized yet
    cov = cli.run(date(2026, 6, 26), sheet_path=sp)
    assert cov.total == 5 and cov.blank == 4 and cov.categorized == 1   # only the manual Health row
    # read-only: the tab is byte-for-byte unchanged (still 4 blank).
    assert sum(1 for v in _by_txn(sp).values() if v == ("", "")) == 4


# ---- review follow-ups (milestone review 2026-06-26) ----

def test_recategorize_reverify_columns_catches_drift(tmp_path):
    """The pre-write header re-validation (§7.1): if the write columns no longer sit
    where the read resolved them (a column shifted between read and write), abort
    rather than stamp the wrong column on the live ledger."""
    sp = _txn_sheet(tmp_path)
    cols = sheet.FINANCE_TRANSACTIONS_COLUMNS
    good = {"Category": cols.index("Category"), "Cat-Source": cols.index("Cat-Source")}
    recat._reverify_columns(good, sp)                       # canonical positions → no raise
    bad = {"Category": 0, "Cat-Source": good["Cat-Source"]}  # col 0 is Date, not Category
    with pytest.raises(recat.RecategorizeError):
        recat._reverify_columns(bad, sp)


def test_coverage_rows_from_grid_normalizes_header_casing():
    """A cased/spaced live header must not make every row read as blank (a false 0%
    on the milestone surface) — rows_from_grid maps headers to canonical keys."""
    grid = [
        ["date", "ACCOUNT", "Description", "category ", "cat-source"],
        ["2026-06-15", "MIZ", "shufersal", "Groceries", "rules"],
    ]
    rows = fincov.rows_from_grid(grid)
    assert rows[0]["Category"] == "Groceries" and rows[0]["Cat-Source"] == "rules"
    assert rows[0]["Account"] == "MIZ"
    c = fincov.coverage(rows)
    assert c.categorized == 1 and c.coverage_pct == 1.0    # NOT a false 0%


def test_excluded_block_is_last_in_rules_file():
    """The Card Settlement exclusion block must be the LAST rules in the file — a
    last-resort fallback below every merchant rule, so a merchant-bearing settlement
    line categorizes by its merchant. Pins the raw file-order invariant directly (the
    existing shadow test checks token concatenation; this checks ordering), since the
    backfill runs the same load_rules/apply_rules and depends on it."""
    rules = categorize.load_rules()
    excluded_idx = [i for i, (_, c) in enumerate(rules) if c in categorize.EXCLUDED_CATEGORIES]
    merchant_idx = [i for i, (_, c) in enumerate(rules) if c not in categorize.EXCLUDED_CATEGORIES]
    assert excluded_idx and merchant_idx                   # both populated
    assert min(excluded_idx) > max(merchant_idx)           # every excluded pattern below every merchant rule
