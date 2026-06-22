"""Seed-safety guard (Lane S / GAP-5).

The committed `Family_OS.xlsx` is both the hermetic test backend AND a public-repo
artifact — it is a binary zip, so the `deploy/publish.sh` `--replace-text` gauntlet
cannot scrub strings inside it. CLAUDE.md's "public-safe by construction" guarantee
therefore rests on the seed carrying NO real PII. This test audits every tab so any
future drift (a real value pasted into the seed) fails before it can be pushed.

Audited 2026-06-18 (all 18 tabs): synthetic by construction — bracketed
placeholders, fake dates, public brand names, `example.com` emails, `000000000`
IDs, sequential last-4s. The ONLY real identifiers are the two adult principals'
first names `Adar`/`Shanee`, which are accepted-public by design: they are the
owner-routing tokens (`OWNER_TO_RECIPIENTS`), the Settings UserMap display names,
and are named throughout CLAUDE.md + the git author identity — so they are NOT
flagged here. High-severity PII (real emails/phones/IDs/JIDs/account numbers) is.
"""
from openpyxl import load_workbook

from automation.lib import config as cfg
from automation.lib import pii   # the shared PII patterns (also back the repo-wide guard)

# The seed must contain at least these tabs — a guard against cfg.SHEET_PATH
# silently pointing at the wrong/empty workbook (a vacuous pass = false safety).
EXPECTED_TABS = {"People", "Contacts", "Finance-Accounts", "Health", "Settings"}


def test_committed_seed_has_no_high_severity_pii():
    wb = load_workbook(cfg.SHEET_PATH, data_only=True)
    missing = EXPECTED_TABS - set(wb.sheetnames)
    assert not missing, (
        f"seed-safety MISCONFIGURED: {cfg.SHEET_PATH} is missing {missing} — this "
        "test would pass vacuously. Point cfg.SHEET_PATH at the real Family_OS seed.")

    scanned, leaks = 0, []
    for tab in wb.sheetnames:
        ws = wb[tab]
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row, col).value
                if v in (None, ""):
                    continue
                scanned += 1
                s, coord = str(v), f"{tab}!{ws.cell(row, col).coordinate}"
                # Identifiers only (pii.scan) — the seed carries synthetic finance
                # amounts by design, so ILS_AMOUNT is NOT applied here.
                for kind in pii.scan(s):
                    leaks.append(f"{coord}: {kind}")

    assert scanned > 50, f"seed-safety MISCONFIGURED: only {scanned} cells scanned in {cfg.SHEET_PATH}"
    assert not leaks, (
        "Family_OS.xlsx (committed → public) carries real PII. The repo is "
        "public-safe by construction (CLAUDE.md) — scrub the offending cell(s) to "
        "placeholders, or restore the clean seed (`git checkout HEAD -- Family_OS.xlsx`), "
        "before any commit/push. Offending cells:\n  " + "\n  ".join(sorted(set(leaks))))
