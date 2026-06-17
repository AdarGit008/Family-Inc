"""Tests for automation/accuracy_review.py — the Phase-F weekly classifier
accuracy review surface (D-048).

Covers: rule re-derivation (reusing the summarizer's hard rules — single source
of truth), trailing-window selection incl. undatable rows, the metric counters,
both renderers, the file-writing run() over an explicit xlsx, the no-live note,
and the weekly-briefing section (present / absent-tab degrade / wiring)."""

import re
from datetime import date

from openpyxl import Workbook

from automation import accuracy_review as ar
from automation.lib import config, sheet


# ---------------------------------------------------------------------------
# Builders
# ---------------------------------------------------------------------------
def _cfg(**overrides):
    base = {
        "group_type": "daycare",
        "importance_default": "alert_eligible",
        "alert_recipients": "both",
        "close_contacts": [],
        "alert_keywords": [],
        "critical_keywords": [],
    }
    base.update(overrides)
    return {"הגן": base}


def _row(**overrides):
    base = {
        "msg_id": "m1", "group_name": "הגן", "group_type": "daycare",
        "sender_name": "הגננת", "sender_role": "unknown",
        "received_at": "2026-06-10T08:00:00", "text": "טקסט כלשהו",
        "has_media": False, "classification": "ALERT", "one_liner": "סיכום קצר",
        "action_required": True, "action_owner": "both", "critical": False,
        "dispatched": True, "dispatched_at": "2026-06-10T08:00:05", "digested_at": "",
    }
    base.update(overrides)
    return base


TODAY = date(2026, 6, 10)  # week window with days=7 → [2026-06-04 .. 2026-06-10]


# ---------------------------------------------------------------------------
# Rule re-derivation — must match the classifier's own hard_rule_alert
# ---------------------------------------------------------------------------
class TestDeriveRule:
    def test_persisted_critical_flag_is_authoritative(self):
        assert ar.derive_rule(_row(critical=True), _cfg()) == ar.RULE_CRITICAL

    def test_critical_flag_as_backend_string(self):
        # gspread RAW round-trips a bool as the string "TRUE"
        assert ar.derive_rule(_row(critical="TRUE"), _cfg()) == ar.RULE_CRITICAL

    def test_alert_keyword_match(self):
        cfg = _cfg(alert_keywords=[re.compile("ביטול")])
        rule = ar.derive_rule(_row(text="שימו לב, ביטול חוגים היום", critical=False), cfg)
        assert "keyword match" in rule

    def test_teacher_evening_window(self):
        rule = ar.derive_rule(
            _row(sender_role="teacher", received_at="2026-06-09T21:30:00",
                 text="מחר להביא כובע", critical=False), _cfg())
        assert rule == "daycare teacher, evening window"

    def test_vaad_utility(self):
        cfg = _cfg(group_type="building")
        rule = ar.derive_rule(
            _row(sender_role="vaad_bayit", text="הפסקת מים מחר", critical=False), cfg)
        assert rule == "vaad bayit utility/maintenance"

    def test_no_hard_rule_is_llm(self):
        rule = ar.derive_rule(
            _row(sender_role="unknown", text="סתם הודעה רגועה", critical=False), _cfg())
        assert rule == ar.RULE_LLM


# ---------------------------------------------------------------------------
# collect() — windowing + counters
# ---------------------------------------------------------------------------
class TestCollect:
    def test_window_includes_boundary_excludes_outside_and_undated(self):
        rows = [
            _row(msg_id="in_today", classification="ALERT", received_at="2026-06-10T07:00:00"),
            _row(msg_id="in_digest", classification="DIGEST", received_at="2026-06-05T09:00:00"),
            _row(msg_id="in_boundary", classification="ROUTINE", received_at="2026-06-04T23:00:00"),
            _row(msg_id="too_old", classification="ALERT", received_at="2026-06-03T09:00:00"),
            _row(msg_id="undated", classification="ALERT", received_at="not-a-date"),
        ]
        m = ar.collect(rows, TODAY, 7, _cfg())
        assert m.start == date(2026, 6, 4) and m.end == TODAY and m.days == 7
        assert m.total_in_window == 3            # today + digest + boundary
        assert m.by_class == {"ALERT": 1, "DIGEST": 1, "ROUTINE": 1}
        assert m.undated == 1                    # excluded from the window
        assert len(m.alerts) == 1

    def test_alert_counters(self):
        rows = [
            _row(msg_id="a1", classification="ALERT", critical=True,
                 action_owner="both", dispatched=True, group_type="daycare",
                 received_at="2026-06-10T07:00:00"),
            _row(msg_id="a2", classification="ALERT", critical=False,
                 action_owner="none", dispatched=False, group_type="family",
                 received_at="2026-06-09T20:00:00"),
        ]
        m = ar.collect(rows, TODAY, 7, _cfg())
        assert len(m.alerts) == 2
        assert m.criticals == 1
        assert m.dispatched == 1
        assert m.floated == 1                    # a2 routed to nobody
        assert m.by_group_type == {"daycare": 1, "family": 1}

    def test_rules_grouped(self):
        cfg = _cfg(alert_keywords=[re.compile("ביטול")])
        rows = [
            _row(msg_id="k1", classification="ALERT", critical=False, text="ביטול היום"),
            _row(msg_id="k2", classification="ALERT", critical=False, text="ביטול מחר"),
            _row(msg_id="c1", classification="ALERT", critical=True, text="חירום"),
        ]
        m = ar.collect(rows, TODAY, 7, cfg)
        assert ar.RULE_CRITICAL in m.by_rule and len(m.by_rule[ar.RULE_CRITICAL]) == 1
        kw = [r for r in m.by_rule if "keyword match" in r]
        assert kw and len(m.by_rule[kw[0]]) == 2

    def test_empty_rows(self):
        m = ar.collect([], TODAY, 7, _cfg())
        assert m.total_in_window == 0 and m.alerts == [] and m.by_rule == {}


# ---------------------------------------------------------------------------
# Renderers
# ---------------------------------------------------------------------------
class TestRender:
    def test_brief_empty_window(self):
        out = ar.render_brief(ar.collect([], TODAY, 7, _cfg()))
        assert "No group messages classified in the last 7 days" in out

    def test_brief_with_alerts_states_counts_rules_and_target(self):
        rows = [_row(classification="ALERT", critical=False, action_owner="none")]
        out = ar.render_brief(ar.collect(rows, TODAY, 7, _cfg()))
        assert "1 ALERT" in out
        assert "ALERTs by rule" in out
        assert "needs a look" in out
        assert f"<{config.ALERT_FP_TARGET_PER_WEEK} ALERT-tier false positive/week" in out

    def test_full_lists_each_alert_grouped_by_rule(self):
        cfg = _cfg(alert_keywords=[re.compile("ביטול")])
        rows = [
            _row(msg_id="a1", classification="ALERT", critical=False,
                 text="ביטול חוגים", one_liner="ביטול חוגים", dispatched=True),
            _row(msg_id="a2", classification="DIGEST", text="סתם"),
        ]
        out = ar.render_full(ar.collect(rows, TODAY, 7, cfg))
        assert "Classifier accuracy review" in out
        assert "ALERTs by triggering rule" in out
        assert "keyword match" in out
        assert "ביטול חוגים" in out
        assert "[sent]" in out
        assert "false positive per week" in out

    def test_full_no_alerts_says_nothing_to_review(self):
        rows = [_row(classification="ROUTINE", one_liner="")]
        out = ar.render_full(ar.collect(rows, TODAY, 7, _cfg()))
        assert "nothing to review" in out
        assert "Target:" in out  # the bar is still stated

    def test_full_flags_undated_and_missing_config(self):
        rows = [
            _row(classification="ALERT", received_at="garbage"),  # undated
            _row(msg_id="m2", classification="ROUTINE", received_at="2026-06-10T09:00:00"),
        ]
        out = ar.render_full(ar.collect(rows, TODAY, 7, {}))  # empty config
        assert "undated, excluded" in out
        assert "group-config seed not found" in out


# ---------------------------------------------------------------------------
# run() — reads the WhatsApp_Inbox tab of an explicit xlsx, writes the surface
# ---------------------------------------------------------------------------
class TestRun:
    def _inbox_xlsx(self, tmp_path, rows):
        p = tmp_path / "s.xlsx"
        wb = Workbook()
        wb.active.title = "README"
        wb.save(p)
        sheet.append_rows(config.WA_INBOX_SHEET_TAB, sheet.WA_INBOX_COLUMNS, rows, p)
        return p

    def test_run_writes_surface_from_sheet(self, tmp_runtime, tmp_path):
        p = self._inbox_xlsx(tmp_path, [
            _row(msg_id="a1", classification="ALERT", critical=False,
                 one_liner="ביטול חוגים", received_at="2026-06-10T08:00:00"),
            _row(msg_id="r1", classification="ROUTINE", received_at="2026-06-09T08:00:00"),
        ])
        out = ar.run(TODAY, 7, sheet_path=p)
        assert out is not None and out.exists()
        body = out.read_text(encoding="utf-8")
        assert "Classifier accuracy review" in body
        assert "ביטול חוגים" in body

    def test_dry_run_writes_nothing(self, tmp_runtime, tmp_path, capsys):
        p = self._inbox_xlsx(tmp_path, [_row(msg_id="a1", classification="ALERT")])
        assert ar.run(TODAY, 7, sheet_path=p, dry_run=True) is None
        assert "Classifier accuracy review" in capsys.readouterr().out

    def test_no_live_backend_says_nothing_to_review(self, tmp_runtime, capsys, monkeypatch):
        class _EmptyWB:
            sheetnames: list = []

            def __getitem__(self, k):
                raise KeyError(k)

        monkeypatch.setattr(ar.sheet, "is_live", lambda: False)
        monkeypatch.setattr(ar.sheet, "workbook", lambda path=None: _EmptyWB())
        ar.run(TODAY, 7, sheet_path=None, dry_run=True)
        assert "no live Sheet backend — nothing to review" in capsys.readouterr().out


# ---------------------------------------------------------------------------
# Weekly-briefing section (Phase F is delivered on the Sat-21:00 cadence)
# ---------------------------------------------------------------------------
class TestBriefingSection:
    def test_absent_inbox_tab_degrades(self):
        from automation.weekly_briefing import section_classifier_accuracy
        wb = Workbook()  # no WhatsApp_Inbox sheet
        assert section_classifier_accuracy(wb, TODAY) == "_No classifier data yet._"

    def test_section_reports_when_tab_present(self):
        from automation.weekly_briefing import section_classifier_accuracy
        wb = Workbook()
        ws = wb.active
        ws.title = config.WA_INBOX_SHEET_TAB
        ws.append(sheet.WA_INBOX_COLUMNS)
        ws.append([_row(msg_id="a1", classification="ALERT",
                        received_at="2026-06-10T08:00:00")[c]
                   for c in sheet.WA_INBOX_COLUMNS])
        out = section_classifier_accuracy(wb, TODAY)
        assert "1 ALERT" in out
        assert "Target" in out

    def test_render_briefing_includes_classifier_section(self, tmp_runtime):
        from automation.weekly_briefing import render_briefing
        wb = Workbook()
        defs = {
            "Calendar-Events": [], "Reminders": [], "Finance-Bdgt": [],
            "Goals": [], "Finance-Accts": [], "People": [],
        }
        for idx, (name, rows) in enumerate(defs.items()):
            ws = wb.active if idx == 0 else wb.create_sheet(name)
            if idx == 0:
                ws.title = name
            for r, row_data in enumerate(rows, start=2):
                for c, val in enumerate(row_data, start=1):
                    ws.cell(r, c, val)
        out = render_briefing(wb, TODAY)
        assert "## Classifier accuracy" in out
        assert "No classifier data yet" in out  # no WhatsApp_Inbox tab → graceful
