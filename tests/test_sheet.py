"""Tests for automation/lib/sheet.py + lib/dates.py — row-parsing tolerance
(ENGINEERING §7: missing columns, bad dates → skipped/None + reported, never
raised), the §7.1 schema-drift guard (headers are strict where rows are
tolerant), the M2 write path, and Settings/UserMap."""

from datetime import date, datetime

import pytest
from openpyxl import Workbook, load_workbook

from automation.lib.dates import add_months, bump_due, fmt_date_he, to_date, to_datetime
from automation.lib.sheet import (
    CellWrite,
    SchemaDriftError,
    _contiguous_runs,
    append_rows,
    encode_value,
    parse_lead_times,
    read_column,
    read_reminders,
    read_settings,
    roll_off_old_rows,
    schema_drift_flag,
    update_reminders,
    validate_reminders_header,
)


# ---------------------------------------------------------------------------
# parse_lead_times
# ---------------------------------------------------------------------------
class TestParseLeadTimes:
    def test_none_defaults(self):
        assert parse_lead_times(None) == [7, 1]

    def test_single_int(self):
        assert parse_lead_times(14) == [14]

    def test_comma_string(self):
        assert parse_lead_times("14, 7, 1") == [14, 7, 1]

    def test_sorts_descending(self):
        assert parse_lead_times("1, 14, 7") == [14, 7, 1]

    def test_filters_junk(self):
        assert parse_lead_times("7, junk, 14") == [14, 7]

    def test_empty_defaults(self):
        assert parse_lead_times("") == [7, 1]


# ---------------------------------------------------------------------------
# to_date
# ---------------------------------------------------------------------------
class TestToDate:
    def test_python_date_passthrough(self):
        d = date(2026, 6, 1)
        assert to_date(d) == d

    def test_datetime_converted(self):
        assert to_date(datetime(2026, 6, 1, 12, 30)) == date(2026, 6, 1)

    def test_iso_string(self):
        assert to_date("2026-06-15") == date(2026, 6, 15)

    def test_sheet_contract_ddmmyyyy(self):
        """SPEC §6.1 col D is DD/MM/YYYY — the gspread port (M2) reads strings."""
        assert to_date("15/06/2026") == date(2026, 6, 15)

    def test_invalid_string_returns_none(self):
        assert to_date("not-a-date") is None

    def test_none_returns_none(self):
        assert to_date(None) is None

    def test_iso_datetime_string(self):
        """The engine stamps Last Sent as ISO datetime TEXT (M2); gspread
        reads it back as a string — to_date must not lose it."""
        assert to_date("2026-06-12T07:30:00") == date(2026, 6, 12)


# ---------------------------------------------------------------------------
# fmt_date_he (DESIGN §6 digest header)
# ---------------------------------------------------------------------------
class TestFmtDateHe:
    def test_friday(self):
        assert fmt_date_he(date(2026, 6, 12)) == "יום ו׳ 12/6"

    def test_sunday_first_weekday(self):
        assert fmt_date_he(date(2026, 6, 14)) == "יום א׳ 14/6"

    def test_none(self):
        assert fmt_date_he(None) == ""


# ---------------------------------------------------------------------------
# bump_due / add_months (SPEC §7.1 recurrence — Feb-29-class clamping)
# ---------------------------------------------------------------------------
class TestBumpDue:
    def test_yearly_plain(self):
        assert bump_due(date(2026, 6, 15), "Yearly") == (date(2027, 6, 15), False)

    def test_weekly(self):
        assert bump_due(date(2026, 6, 15), "Weekly") == (date(2026, 6, 22), False)

    def test_monthly_jan31_clamps_to_feb_end(self):
        assert bump_due(date(2026, 1, 31), "Monthly") == (date(2026, 2, 28), True)

    def test_feb29_yearly_clamps_and_flags(self):
        assert bump_due(date(2028, 2, 29), "Yearly") == (date(2029, 2, 28), True)

    def test_feb29_into_leap_year_keeps_day(self):
        assert add_months(date(2027, 2, 28), 12) == (date(2028, 2, 28), False)

    def test_quarterly(self):
        assert bump_due(date(2026, 11, 30), "Quarterly") == (date(2027, 2, 28), True)

    def test_one_off_and_custom_unbumpable(self):
        assert bump_due(date(2026, 6, 15), "One-off") == (None, False)
        assert bump_due(date(2026, 6, 15), "Custom") == (None, False)
        assert bump_due(date(2026, 6, 15), "every full moon") == (None, False)


# ---------------------------------------------------------------------------
# to_datetime (dashboard write-back shapes)
# ---------------------------------------------------------------------------
class TestToDatetime:
    def test_iso_with_t(self):
        assert to_datetime("2026-06-10T09:30:00") == datetime(2026, 6, 10, 9, 30)

    def test_space_separator(self):
        assert to_datetime("2026-06-10 09:30:00") == datetime(2026, 6, 10, 9, 30)

    def test_date_only_string(self):
        assert to_datetime("2026-06-10") == datetime(2026, 6, 10)

    def test_garbage_returns_none(self):
        assert to_datetime("soonish") is None

    def test_empty_returns_none(self):
        assert to_datetime("  ") is None
        assert to_datetime(None) is None


# ---------------------------------------------------------------------------
# read_reminders — tolerance against real-world sheet mess
# ---------------------------------------------------------------------------
def _write_sheet(tmp_path, rows):
    """rows = list of value-lists laid into Reminders!A2:P…"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reminders"
    ws.append(["Title", "Domain", "Owner", "Due Date", "Lead Times", "Recurrence",
               "Status", "Last Sent", "Channel", "Notes", "Days Until", "Auto-flag",
               "LastDoneBy", "DoneAt", "WriteQueue_Tombstone", "Guide URL"])
    for row in rows:
        ws.append(row)
    p = tmp_path / "sheet.xlsx"
    wb.save(p)
    return p


class TestReadReminders:
    def test_basic_row(self, tmp_path):
        p = _write_sheet(tmp_path, [
            ["Car test", "Car", "Both", date(2026, 7, 1), "14,7", "Yearly", "Pending"],
        ])
        out = read_reminders(p)
        assert len(out) == 1
        r = out[0]
        assert r.title == "Car test"
        assert r.due == date(2026, 7, 1)
        assert r.lead_times == [14, 7]
        assert r.owner == "Both"

    def test_blank_and_templated_rows_skipped(self, tmp_path):
        p = _write_sheet(tmp_path, [
            [None, "Car", "Adar", date(2026, 7, 1)],
            ["[Template] copy me", "Car", "Adar", date(2026, 7, 1)],
            ["Real", "Car", "Adar", date(2026, 7, 1)],
        ])
        out = read_reminders(p)
        assert [r.title for r in out] == ["Real"]

    def test_bad_date_becomes_none_not_raise(self, tmp_path):
        p = _write_sheet(tmp_path, [
            ["Fuzzy due", "Other", "Adar", "sometime in July"],
        ])
        out = read_reminders(p)
        assert len(out) == 1
        assert out[0].due is None  # engine classifies it out; hygiene reports it

    def test_missing_mno_columns_ok(self, tmp_path):
        """Legacy rows without LastDoneBy/DoneAt/Tombstone parse fine
        (SPEC §9: additive-only schema — old rows keep parsing)."""
        p = _write_sheet(tmp_path, [
            ["Old row", "Health", "Shanee", date(2026, 8, 1), "7"],
        ])
        r = read_reminders(p)[0]
        assert r.last_done_by == ""
        assert r.done_at is None
        assert r.write_queue_tombstone is None

    def test_defaults_for_empty_cells(self, tmp_path):
        p = _write_sheet(tmp_path, [
            ["Bare", None, None, date(2026, 8, 1)],
        ])
        r = read_reminders(p)[0]
        assert r.domain == "Other"
        assert r.owner == "Adar"
        assert r.status == "Pending"
        assert r.channel == "WhatsApp"
        assert r.lead_times == [7, 1]


# ---------------------------------------------------------------------------
# Schema-drift guard (SPEC §7.1) — headers strict, abort + flag on mismatch
# ---------------------------------------------------------------------------
def _write_sheet_with_header(tmp_path, header, rows=()):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reminders"
    ws.append(header)
    for row in rows:
        ws.append(row)
    p = tmp_path / "drift.xlsx"
    wb.save(p)
    return p


GOOD_HEADER = ["Title", "Domain", "Owner", "Due Date", "Lead Times", "Recurrence",
               "Status", "Last Sent", "Channel", "Notes", "Days Until", "Auto-flag",
               "LastDoneBy", "DoneAt", "WriteQueue_Tombstone", "Guide URL"]


class TestSchemaGuard:
    def test_clean_header_passes(self):
        assert validate_reminders_header(GOOD_HEADER) == []

    def test_case_and_whitespace_tolerated(self):
        relaxed = [h.upper() + " " for h in GOOD_HEADER]
        assert validate_reminders_header(relaxed) == []

    def test_extra_columns_beyond_p_tolerated(self):
        assert validate_reminders_header(GOOD_HEADER + ["Anything"]) == []

    def test_renamed_column_reported(self):
        bad = GOOD_HEADER.copy()
        bad[4] = "Lead Times (days)"  # the pre-M2 as-built drift
        problems = validate_reminders_header(bad)
        assert len(problems) == 1 and "col 5" in problems[0]

    def test_missing_trailing_columns_reported(self):
        problems = validate_reminders_header(GOOD_HEADER[:12])
        assert len(problems) == 4  # M, N, O, P absent

    def test_drifted_read_aborts_and_flags(self, tmp_runtime, tmp_path):
        bad = GOOD_HEADER.copy()
        bad[6] = "State"
        p = _write_sheet_with_header(tmp_path, bad)
        with pytest.raises(SchemaDriftError):
            read_reminders(p)
        flag = schema_drift_flag()
        assert flag and any("col 7" in x for x in flag["problems"])

    def test_drifted_write_refused(self, tmp_runtime, tmp_path):
        """The guard protects BOTH directions of the dual write path."""
        bad = GOOD_HEADER.copy()
        bad[7] = "LastSent"
        p = _write_sheet_with_header(tmp_path, bad)
        with pytest.raises(SchemaDriftError):
            update_reminders([CellWrite(2, "Status", "Sent")], p)

    def test_clean_read_heals_the_flag(self, tmp_runtime, tmp_path):
        bad = _write_sheet_with_header(tmp_path, GOOD_HEADER[:10])
        with pytest.raises(SchemaDriftError):
            read_reminders(bad)
        assert schema_drift_flag() is not None
        good = _write_sheet(tmp_path, [["OK", "Car", "Adar", date(2026, 7, 1)]])
        read_reminders(good)
        assert schema_drift_flag() is None


# ---------------------------------------------------------------------------
# Write path — update_reminders / append_rows / read_column (xlsx backend;
# the gspread backend shares encode_value and the same call shapes)
# ---------------------------------------------------------------------------
class TestEncodeValue:
    def test_date_is_iso(self):
        # Lane C: col-D (the only date-typed write) emits ISO YYYY-MM-DD — Sheets
        # parses ISO locale-unambiguously; the dashboard's parseDate reads back
        # either ISO or the he-IL DD/MM render. round-trips via to_date below.
        assert encode_value(date(2028, 2, 29)) == "2028-02-29"

    def test_due_date_round_trips_through_to_date(self):
        from automation.lib.dates import to_date
        assert to_date(encode_value(date(2026, 6, 25))) == date(2026, 6, 25)

    def test_datetime_is_iso_t_text(self):
        assert encode_value(datetime(2026, 6, 12, 7, 30)) == "2026-06-12T07:30:00"

    def test_none_clears(self):
        assert encode_value(None) == ""

    def test_bool_uppercase(self):
        assert encode_value(True) == "TRUE"


class TestUpdateReminders:
    def test_batch_lands_and_roundtrips(self, tmp_path):
        p = _write_sheet(tmp_path, [
            ["Car test", "Car", "Both", date(2026, 7, 1), "14,7", "Yearly", "Pending"],
        ])
        now = datetime(2026, 6, 12, 7, 31, 2)
        update_reminders([
            CellWrite(2, "Last Sent", now),
            CellWrite(2, "Status", "Sent"),
        ], p)
        r = read_reminders(p)[0]
        assert r.status == "Sent"
        assert r.last_sent == date(2026, 6, 12)

    def test_clear_with_none(self, tmp_path):
        p = _write_sheet(tmp_path, [
            ["Car test", "Car", "Both", date(2026, 7, 1), "14,7", "Yearly", "Sent",
             datetime(2026, 6, 1, 7, 30)],
        ])
        update_reminders([CellWrite(2, "Last Sent", None)], p)
        assert read_reminders(p)[0].last_sent is None

    def test_unknown_field_rejected(self):
        with pytest.raises(ValueError):
            CellWrite(2, "Pet Name", "x")

    def test_header_row_protected(self):
        with pytest.raises(ValueError):
            CellWrite(1, "Status", "x")

    def test_formulas_survive_a_write(self, tmp_path):
        """Reads are data_only, writes must NOT be — saving a data_only
        workbook silently deletes every formula (cols K/L are formulas on the
        real sheet)."""
        p = _write_sheet(tmp_path, [
            ["Car test", "Car", "Both", date(2026, 7, 1), "14,7", "Yearly", "Pending"],
        ])
        wb = load_workbook(p)
        wb["Reminders"].cell(2, 11).value = "=D2-TODAY()"
        wb.save(p)
        update_reminders([CellWrite(2, "Status", "Sent")], p)
        assert load_workbook(p)["Reminders"].cell(2, 11).value == "=D2-TODAY()"


class TestAppendAndReadColumn:
    def test_append_creates_tab_with_header(self, tmp_path):
        p = _write_sheet(tmp_path, [])
        append_rows("WhatsApp_Inbox", ["msg_id", "text"],
                    [{"msg_id": "m1", "text": "שלום"}], p)
        append_rows("WhatsApp_Inbox", ["msg_id", "text"],
                    [{"msg_id": "m2", "text": "עוד"}], p)
        assert read_column("WhatsApp_Inbox", "msg_id", p) == ["m1", "m2"]

    def test_read_column_missing_tab_or_column_degrades(self, tmp_path):
        p = _write_sheet(tmp_path, [])
        assert read_column("Nope", "msg_id", p) == []
        append_rows("T", ["a"], [{"a": 1}], p)
        assert read_column("T", "missing", p) == []


# ---------------------------------------------------------------------------
# Settings (SPEC §6.4): UserMap + lang
# ---------------------------------------------------------------------------
class TestSettings:
    def test_usermap_and_lang(self, tmp_path):
        p = _write_sheet(tmp_path, [])
        append_rows("Settings", ["Key", "Value"], [
            {"Key": "You@Example.com", "Value": "Adar"},
            {"Key": "partner@example.com", "Value": "Shanee"},
            {"Key": "lang", "Value": "he"},
        ], p)
        s = read_settings(p)
        assert s.lang == "he"
        assert s.display_name("you@example.com") == "Adar"   # case-folded
        assert s.display_name("partner@example.com") == "Shanee"
        assert s.display_name("stranger@example.com", "You") == "You"

    def test_missing_tab_defaults(self, tmp_path):
        p = _write_sheet(tmp_path, [])
        s = read_settings(p)
        assert s.lang == "he" and s.usermap == {}


# ---------------------------------------------------------------------------
# Hot-tab rolloff (SPEC §6.2, M4/D-044): WhatsApp_Inbox keeps a 30-day window;
# delete-by-date, header kept, undatable rows kept, seed never touched
# ---------------------------------------------------------------------------
def _write_wa_inbox(tmp_path, rows):
    p = _write_sheet(tmp_path, [])  # gives a workbook with a Reminders tab
    append_rows("WhatsApp_Inbox", ["msg_id", "received_at", "text"],
                [{"msg_id": m, "received_at": r, "text": t} for m, r, t in rows], p)
    return p


class TestContiguousRuns:
    def test_collapses_runs(self):
        assert _contiguous_runs([2, 3, 4, 7, 9, 10]) == [(2, 4), (7, 7), (9, 10)]

    def test_empty(self):
        assert _contiguous_runs([]) == []


class TestRollOff:
    def test_deletes_only_rows_before_cutoff(self, tmp_path):
        p = _write_wa_inbox(tmp_path, [
            ("m1", "2026-05-01T08:00:00", "old"),
            ("m2", "2026-05-20T08:00:00", "old2"),
            ("m3", "2026-06-10T08:00:00", "keep"),
        ])
        assert roll_off_old_rows("WhatsApp_Inbox", "received_at", date(2026, 5, 25), p) == 2
        assert read_column("WhatsApp_Inbox", "msg_id", p) == ["m3"]

    def test_keeps_blank_and_unparseable_dates(self, tmp_path):
        p = _write_wa_inbox(tmp_path, [
            ("m1", "", "no date"),
            ("m2", "garbage", "bad date"),
            ("m3", "2026-01-01T00:00:00", "old"),
        ])
        assert roll_off_old_rows("WhatsApp_Inbox", "received_at", date(2026, 6, 1), p) == 1
        assert set(read_column("WhatsApp_Inbox", "msg_id", p)) == {"m1", "m2"}

    def test_scattered_old_rows_delete_correctly(self, tmp_path):
        p = _write_wa_inbox(tmp_path, [
            ("m1", "2026-01-01T00:00:00", "old"),
            ("m2", "2026-06-10T00:00:00", "new"),
            ("m3", "2026-01-02T00:00:00", "old"),
            ("m4", "2026-06-11T00:00:00", "new"),
        ])
        assert roll_off_old_rows("WhatsApp_Inbox", "received_at", date(2026, 6, 1), p) == 2
        assert read_column("WhatsApp_Inbox", "msg_id", p) == ["m2", "m4"]

    def test_missing_tab_or_column_is_noop(self, tmp_path):
        p = _write_wa_inbox(tmp_path, [("m1", "2026-01-01T00:00:00", "x")])
        assert roll_off_old_rows("Nope", "received_at", date(2026, 6, 1), p) == 0
        assert roll_off_old_rows("WhatsApp_Inbox", "nope_col", date(2026, 6, 1), p) == 0

    def test_skips_without_live_backend_or_path(self):
        """No explicit path + not live (conftest blanks SHEET_ID) → 0, and the
        committed seed is never opened (same posture as the write-backs)."""
        assert roll_off_old_rows("WhatsApp_Inbox", "received_at", date(2026, 6, 1)) == 0
