"""Tests for sunday_briefing.py — section builders with mock workbook data."""

from datetime import date, timedelta

import pytest
from openpyxl import Workbook

from sunday_briefing import (
    section_week_ahead,
    section_reminders_week,
    section_money,
    section_goals,
    section_hygiene,
    render_briefing,
    WEEK_AHEAD_DAYS,
)


def _make_wb(sheet_defs: dict) -> Workbook:
    """Create an in-memory openpyxl Workbook with named sheets populated from
    list-of-lists data (row 1 = header placeholders so row 2+ is data)."""
    wb = Workbook()
    for idx, (sheet_name, rows) in enumerate(sheet_defs.items()):
        if idx == 0:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(sheet_name)
        for r, row_data in enumerate(rows, start=2):
            for c, val in enumerate(row_data, start=1):
                ws.cell(r, c, val)
    return wb


# ---------------------------------------------------------------------------
# section_week_ahead
# ---------------------------------------------------------------------------
class TestSectionWeekAhead:
    def test_empty_when_no_events(self):
        wb = _make_wb({"Calendar-Events": []})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Quiet week" in result

    def test_shows_events_in_range(self):
        events = [
            [date(2026, 6, 12), "10:00", "11:00", "Dentist", "Adar", "", "Clinic"],
            [date(2026, 6, 14), "", "", "Family lunch", "Both", "", "Home"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Dentist" in result
        assert "Family lunch" in result
        assert "10:00–11:00" in result

    def test_excludes_events_before_today(self):
        events = [
            [date(2026, 6, 8), "09:00", "10:00", "Past event", "Adar"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Past event" not in result

    def test_excludes_events_after_window(self):
        events = [
            [date(2026, 6, 20), "09:00", "10:00", "Future event", "Adar"],
        ]
        wb = _make_wb({"Calendar-Events": events})
        today = date(2026, 6, 10)
        result = section_week_ahead(wb, today, today + timedelta(days=7))
        assert "Future event" not in result


# ---------------------------------------------------------------------------
# section_reminders_week
# ---------------------------------------------------------------------------
class TestSectionRemindersWeek:
    def test_empty_when_no_reminders(self):
        wb = _make_wb({"Reminders": []})
        today = date(2026, 6, 10)
        upcoming, overdue = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "No upcoming" in upcoming
        assert "No overdue" in overdue

    def test_upcoming_within_week(self):
        reminders = [
            ["Dentist checkup", "Health", "Adar", date(2026, 6, 14), "14,7,1", "One-off", "Pending"],
            ["Car test", "Auto", "Adar", date(2026, 6, 12), "7,1", "Annual", "Pending"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Dentist checkup" in upcoming
        assert "Car test" in upcoming

    def test_overdue_items(self):
        reminders = [
            ["Late bill", "Finance", "Adar", date(2026, 6, 1), "7,1", "Monthly", "Pending"],
            ["Future", "Kids", "Adar", date(2026, 6, 20), "7,1", "One-off", "Pending"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        _, overdue = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Late bill" in overdue
        assert "overdue" in overdue.lower()

    def test_skips_done_and_skipped(self):
        reminders = [
            ["Done item", "Kids", "Adar", date(2026, 6, 12), "7,1", "One-off", "Done"],
            ["Skipped item", "Kids", "Adar", date(2026, 6, 12), "7,1", "One-off", "Skipped"],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Done item" not in upcoming
        assert "Skipped item" not in upcoming

    def test_skips_templated_rows(self):
        reminders = [
            ["[Template] thing", "Kids", "Adar", date(2026, 6, 12)],
        ]
        wb = _make_wb({"Reminders": reminders})
        today = date(2026, 6, 10)
        upcoming, _ = section_reminders_week(wb, today, today + timedelta(days=7))
        assert "Template" not in upcoming


# ---------------------------------------------------------------------------
# section_money
# ---------------------------------------------------------------------------
class TestSectionMoney:
    def test_basic_totals(self):
        rows = [
            ["Groceries", 3000, 2500],
            ["Rent", 5000, 5000],
            ["Utilities", 1000, 1200],
        ]
        wb = _make_wb({"Finance-Bdgt": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "₪8,700" in result  # 2500+5000+1200
        assert "₪9,000" in result  # 3000+5000+1000
        assert "Utilities" in result  # over-budget category
        assert "Groceries" not in result  # under budget, not in top 3

    def test_no_over_budget(self):
        rows = [
            ["Groceries", 3000, 2000],
            ["Rent", 5000, 5000],
        ]
        wb = _make_wb({"Finance-Bdgt": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "No categories over budget" in result

    def test_skips_total_row(self):
        rows = [
            ["Groceries", 3000, 2500],
            ["TOTAL", 10000, 9000],
        ]
        wb = _make_wb({"Finance-Bdgt": rows})
        today = date(2026, 6, 10)
        result = section_money(wb, today)
        assert "TOTAL" not in result


# ---------------------------------------------------------------------------
# section_goals
# ---------------------------------------------------------------------------
class TestSectionGoals:
    def test_empty_goals(self):
        wb = _make_wb({"Goals": []})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "No goals tracked" in result

    def test_goal_with_milestone_flag(self):
        rows = [
            ["Lose 5kg", "Shanee", "", date(2026, 6, 25), "Start walking 3x/week", 40, date(2026, 6, 5), "In progress"],
        ]
        wb = _make_wb({"Goals": rows})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "Lose 5kg" in result
        assert "milestone in 15d" in result

    def test_stale_goal_warning(self):
        rows = [
            ["Read books", "Adar", "", date(2026, 12, 31), "", 10, date(2026, 5, 1), "In progress"],
        ]
        wb = _make_wb({"Goals": rows})
        today = date(2026, 6, 10)
        result = section_goals(wb, today)
        assert "no update" in result


# ---------------------------------------------------------------------------
# section_hygiene
# ---------------------------------------------------------------------------
class TestSectionHygiene:
    def test_all_clean(self):
        wb = _make_wb({
            "Reminders": [["OK", "", "", date(2026, 6, 15)]],
            "Finance-Accts": [["Bank", "", "", "", "", "", date(2026, 6, 1)]],
            "People": [["Adar"]],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "All clean" in result

    def test_missing_due_date(self):
        wb = _make_wb({
            "Reminders": [["Missing due", "", "", None]],
            "Finance-Accts": [],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "missing a due date" in result.lower()

    def test_stale_account(self):
        wb = _make_wb({
            "Reminders": [],
            "Finance-Accts": [["Old Bank", "", "", "", "", "", date(2026, 1, 1)]],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "not imported in" in result

    def test_placeholder_people(self):
        wb = _make_wb({
            "Reminders": [],
            "Finance-Accts": [],
            "People": [["[replace me]"]],
        })
        today = date(2026, 6, 10)
        result = section_hygiene(wb, today)
        assert "placeholder" in result


# ---------------------------------------------------------------------------
# render_briefing (integration smoke)
# ---------------------------------------------------------------------------
class TestRenderBriefing:
    def test_renders_all_sections(self):
        wb = _make_wb({
            "Calendar-Events": [[date(2026, 6, 12), "10:00", "11:00", "Dentist", "Adar"]],
            "Reminders": [["Upcoming", "Health", "Adar", date(2026, 6, 14), "7", "One-off", "Pending"]],
            "Finance-Bdgt": [["Groceries", 3000, 2500]],
            "Goals": [["Goal 1", "Adar", "", None, "", 0, None, "Not started"]],
            "Finance-Accts": [],
            "People": [],
        })
        today = date(2026, 6, 10)
        result = render_briefing(wb, today)
        assert "Family inc. — Sunday Briefing" in result
        assert "Week ahead" in result
        assert "Reminders firing this week" in result
        assert "Overdue" in result
        assert "Money" in result
        assert "Goals" in result
        assert "Data hygiene" in result
