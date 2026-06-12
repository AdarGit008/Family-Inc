"""
Family inc. — Reply handler for reminder WhatsApp replies.

PARKED until v1.1 (BACKLOG.md: reply parsing — needs the bridge's 1:1 read
guard lifted for exactly the two adult JIDs, plus tests). Not wired to any
timer. Reinstatement also ports the sheet writes to lib/sheet (gspread).

Outbox posture (M2, D-015): everything goes through lib/outbox.queue() with
stable ids wa-{msg_id} (idempotent per inbound reply). Kinds as wired today:
acks/errors = "alert", digest re-send = "briefing". OPEN for the v1.1
reinstatement PO call: a solicited reply-ack arguably shouldn't consume the
2/day unsolicited budget or hold for quiet hours — that likely wants a new
outbox kind, which is a SPEC §7.5 contract change and a milestone review.

Reads bridge/state/inbox/replies.jsonl (written by the Baileys bridge when
Adar/Shanee reply to reminder digests), applies the commands to the
Reminders tab, and sends follow-up responses via the outbox queue.

Commands:
  done, 1 done, 1 ✅          → set Status = Done
  +7, 1 +7, snooze 7d         → push Due Date by N days, Status = Snoozed
  mute 30d, 1 mute            → Status = Snoozed for 30d; doesn't change Due Date
  list, today, ?              → send current digest re-rendered

Unrecognized commands receive a one-time "didn't catch that" reply (the
bridge sends this immediately; the engine only logs unrecognized commands
and skips them).

Run modes:
  python reply_handler.py              # process all unprocessed replies
  python reply_handler.py --dry-run    # print actions, no sheet changes
"""

from __future__ import annotations

if __package__ in (None, ""):  # direct `python3 automation/reply_handler.py`
    import sys as _sys
    from pathlib import Path as _Path
    _sys.path.insert(0, str(_Path(__file__).resolve().parent.parent))

import argparse
import json
import logging
import re
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook

from automation.lib import config
from automation.lib.outbox import queue

SHEET_PATH = config.SHEET_PATH
REPLIES_FILE = config.BRIDGE_STATE_DIR / "inbox" / "replies.jsonl"
PROCESSED_FILE = config.BRIDGE_STATE_DIR / "inbox" / "replies_processed.jsonl"

log = logging.getLogger("reply_handler")

# Regex for parsing reply commands (mirrors bridge-side parsing for defense in depth)
RE_DONE = re.compile(r"^(?:done|✅)\s*$", re.IGNORECASE)
RE_SNOOZE = re.compile(r"^\+(\d+)\s*$|^snooze\s+(\d+)d?\s*$", re.IGNORECASE)
RE_MUTE = re.compile(r"^mute\s*$|^mute\s+(\d+)d?\s*$", re.IGNORECASE)
RE_LIST = re.compile(r"^(?:list|today|\?)\s*$", re.IGNORECASE)


def _parse_index(text: str) -> tuple[Optional[int], str]:
    """Extract optional numeric index prefix (e.g., '2 +7' -> (2, '+7'))."""
    m = re.match(r"^(\d+)\s+(.+)$", text.strip())
    if m:
        return int(m.group(1)), m.group(2)
    return None, text.strip()


def _load_processed_ids() -> set[str]:
    """msg_ids already handled, so reruns are idempotent."""
    if not PROCESSED_FILE.exists():
        return set()
    ids = set()
    for line in PROCESSED_FILE.read_text(encoding="utf-8").splitlines():
        try:
            row = json.loads(line)
            if row.get("msg_id"):
                ids.add(row["msg_id"])
        except (json.JSONDecodeError, KeyError):
            continue
    return ids


def _mark_processed(msg_id: str, action: str, detail: str) -> None:
    """Append a processed entry so this reply is never re-applied."""
    row = {
        "msg_id": msg_id,
        "action": action,
        "detail": detail,
        "processed_at": datetime.now().astimezone().isoformat(timespec="seconds"),
    }
    with PROCESSED_FILE.open("a", encoding="utf-8") as fh:
        fh.write(json.dumps(row, ensure_ascii=False) + "\n")


def read_digest_for_today() -> Optional[str]:
    """Read the most recent reminders briefing for today's date.
    Returns the markdown body or None if no briefing exists."""
    today_iso = date.today().isoformat()
    briefings_dir = config.BRIEFINGS_DIR
    if not briefings_dir.exists():
        return None
    # Find today's briefing files (briefing_<recipient>.md)
    candidates = sorted(
        briefings_dir.glob(f"{today_iso}_briefing_*.md"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    if not candidates:
        return None
    return candidates[0].read_text(encoding="utf-8").strip()


def _find_reminder_row(ws, index: int) -> Optional[int]:
    """Map a 1-based display index to a sheet row (2+). The index is the
    position in the digest display, not the sheet row number."""
    count = 0
    for row_idx in range(2, ws.max_row + 1):
        title = ws.cell(row_idx, 1).value
        if not title or str(title).startswith("["):
            continue
        count += 1
        if count == index:
            return row_idx
    return None


def apply_done(ws, row_idx: int) -> str:
    """Mark a reminder row as Done."""
    ws.cell(row_idx, 7).value = "Done"
    ws.cell(row_idx, 13).value = "Adar"  # last_done_by
    ws.cell(row_idx, 14).value = datetime.now()  # done_at
    title = ws.cell(row_idx, 1).value or "item"
    return f"Marked '{title}' (row {row_idx}) as Done"


def apply_snooze(ws, row_idx: int, days: int) -> str:
    """Push Due Date by N days, set Status = Snoozed."""
    current_due = ws.cell(row_idx, 4).value
    if isinstance(current_due, datetime):
        current_due = current_due.date()
    elif isinstance(current_due, str):
        try:
            current_due = datetime.strptime(current_due.strip(), "%Y-%m-%d").date()
        except ValueError:
            current_due = date.today()
    if not isinstance(current_due, date):
        current_due = date.today()

    new_due = current_due + timedelta(days=days)
    ws.cell(row_idx, 4).value = new_due
    ws.cell(row_idx, 7).value = "Snoozed"
    title = ws.cell(row_idx, 1).value or "item"
    return f"Snoozed '{title}' (row {row_idx}) by {days} days → due {new_due.isoformat()}"


def apply_mute(ws, row_idx: int, days: int) -> str:
    """Set Status = Snoozed for N days without changing Due Date."""
    ws.cell(row_idx, 7).value = "Snoozed"
    title = ws.cell(row_idx, 1).value or "item"
    return f"Muted '{title}' (row {row_idx}) for {days} days"


def process_replies(dry_run: bool = False) -> dict:
    """Read replies.jsonl, apply commands to the Reminders sheet, send responses.
    Returns summary dict."""
    if not REPLIES_FILE.exists():
        log.info("No replies to process.")
        return {"processed": 0, "actions": []}

    processed_ids = _load_processed_ids()
    wb = None
    actions = []
    processed_count = 0

    replies = []
    for line in REPLIES_FILE.read_text(encoding="utf-8").splitlines():
        try:
            replies.append(json.loads(line))
        except json.JSONDecodeError:
            continue

    for reply in replies:
        msg_id = reply.get("msg_id", "")
        if not msg_id or msg_id in processed_ids:
            continue

        text = reply.get("text", "").strip()
        parsed = reply.get("parsed")  # pre-parsed by bridge

        if not parsed or not parsed.get("cmd"):
            # Unrecognized command — bridge already sent help text; skip
            _mark_processed(msg_id, "unrecognized", text[:80])
            processed_count += 1
            continue

        cmd = parsed["cmd"]
        index = parsed.get("index")
        n = parsed.get("n")

        # Commands that need sheet access
        if cmd in ("done", "snooze", "mute"):
            if wb is None:
                wb = load_workbook(SHEET_PATH)
            ws = wb["Reminders"]

            if index is not None:
                row_idx = _find_reminder_row(ws, index)
                if row_idx is None:
                    queue(
                        "both",
                        f"❓ Couldn't find reminder #{index} in today's list. "
                        "Reply ? to see the current list.",
                        "alert",
                        source="reply_handler", msg_id=f"wa-{msg_id}",
                    )
                    _mark_processed(msg_id, cmd, f"index {index} not found")
                    processed_count += 1
                    continue
            else:
                # No index specified — apply to the first non-done reminder in
                # today's digest order
                row_idx = _find_reminder_row(ws, 1)
                if row_idx is None:
                    queue(
                        "both",
                        "❓ No active reminders found. Reply ? to check.",
                        "alert",
                        source="reply_handler", msg_id=f"wa-{msg_id}",
                    )
                    _mark_processed(msg_id, cmd, "no active reminders")
                    processed_count += 1
                    continue

            if cmd == "done":
                detail = apply_done(ws, row_idx)
                actions.append(detail)
                if not dry_run:
                    queue("both", f"✅ {detail}", "alert",
                          source="reply_handler", msg_id=f"wa-{msg_id}")
            elif cmd == "snooze":
                detail = apply_snooze(ws, row_idx, n or 7)
                actions.append(detail)
                if not dry_run:
                    queue("both", f"📆 {detail}", "alert",
                          source="reply_handler", msg_id=f"wa-{msg_id}")
            elif cmd == "mute":
                detail = apply_mute(ws, row_idx, n or 30)
                actions.append(detail)
                if not dry_run:
                    queue("both", f"🤐 {detail}", "alert",
                          source="reply_handler", msg_id=f"wa-{msg_id}")

        elif cmd in ("list", "today", "?"):
            digest = read_digest_for_today()
            if digest:
                actions.append("Sent today's digest on request")
                if not dry_run:
                    # A solicited re-send of the briefing — kind=briefing
                    # (budget-exempt; SPEC §7.5).
                    queue("both", digest, "briefing",
                          source="reply_handler", msg_id=f"wa-{msg_id}")
            else:
                actions.append("No digest available for today")
                if not dry_run:
                    queue("both",
                          "📋 No reminders digest for today yet. "
                          "The engine runs at 07:30 daily.",
                          "alert",
                          source="reply_handler", msg_id=f"wa-{msg_id}")

        elif cmd == "help":
            # Bridge already sent help; nothing more to do
            actions.append("Help text already sent by bridge")
            pass

        _mark_processed(msg_id, cmd, actions[-1] if actions else "")
        processed_count += 1

    # Save sheet changes
    if wb is not None and not dry_run and any(
        a for a in actions if not a.startswith("Sent") and not a.startswith("Help") and not a.startswith("No digest")
    ):
        wb.save(SHEET_PATH)
        log.info("Saved changes to %s", SHEET_PATH)

    if dry_run:
        print("[DRY RUN] Would apply:")
        for a in actions:
            print(f"  - {a}")

    return {"processed": processed_count, "actions": actions}


def main():
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="Print actions, no sheet changes")
    args = ap.parse_args()

    result = process_replies(dry_run=args.dry_run)
    print(f"Processed {result['processed']} replies")
    for a in result["actions"]:
        print(f"  {a}")


if __name__ == "__main__":
    main()
