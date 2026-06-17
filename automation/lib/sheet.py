"""
Family inc. — master-DB access layer. The ONLY module that opens the workbook
(D-016; ENGINEERING.md §1: nothing outside this file constructs a gspread
client).

Two backends behind one surface:

  live (gspread + service account) — active when FAMILY_INC_SHEET_ID is set
      (the appliance; creds at /etc/family-inc/service-account.json or
      $FAMILY_INC_SA_JSON). Missing creds with the id set = hard error, not a
      silent fallback (fail loud).
  seed (openpyxl xlsx)             — everything else: tests and creds-less dev.
      Reads default to the committed seed template; WRITES never touch the
      seed by default — write paths take an explicit path (tests use tmp
      copies) and callers gate on `is_live()` (see reminders_engine.run).

Public surface:
  Reminder / read_reminders()        — Reminders tab → dataclasses, schema-guarded
  CellWrite / update_reminders()     — one batched write-back, schema-guarded
  append_rows()                      — append dict-rows to a tab (creates tab+header)
  Settings / read_settings()         — UserMap (email → display name) + lang
  workbook()                         — duck-typed read handle (tab["Name"].cell(r,c).value)
                                       for the loose readers (weekly briefing)
  SchemaDriftError / is_live()

Schema-drift guard (SPEC.md §7.1): every Reminders read/write first validates
the header row against the §6.1 column map. Mismatch → SchemaDriftError +
logs/schema_drift.flag (the weekly briefing surfaces it; a clean read clears
it). This guards the dual write path — dashboard and engine must agree on
columns before anything fires.

Row-parsing posture (SPEC.md §6 + §9): tolerate missing columns, return None
for bad dates, skip blank/templated rows. Bad data is skipped + reported in
data-hygiene lines — it never raises. Headers are strict; rows are tolerant.
"""
from __future__ import annotations

import json
import logging
import os
import time as _time
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from automation.lib import config
from automation.lib.dates import to_date, to_datetime

log = logging.getLogger("sheet")

# SPEC.md §6.1 column map (1-based). The single source for both the parser
# and the schema-drift guard; the dashboard writes by these positions too.
REMINDERS_COLUMNS = {
    1: "Title", 2: "Domain", 3: "Owner", 4: "Due Date", 5: "Lead Times",
    6: "Recurrence", 7: "Status", 8: "Last Sent", 9: "Channel", 10: "Notes",
    11: "Days Until", 12: "Auto-flag", 13: "LastDoneBy", 14: "DoneAt",
    15: "WriteQueue_Tombstone", 16: "Guide URL",
}
FIELD_TO_COL = {name: col for col, name in REMINDERS_COLUMNS.items()}

WA_INBOX_COLUMNS = [
    "msg_id", "group_name", "group_type", "sender_name", "sender_role",
    "received_at", "text", "has_media", "classification", "one_liner",
    "action_required", "action_owner", "critical", "dispatched",
    "dispatched_at", "digested_at",
]
WA_ARCHIVE_COLUMNS = ["msg_id", "group_name", "sender_name", "received_at",
                      "text", "one_liner"]
SETTINGS_COLUMNS = ["Key", "Value"]

# SPEC §12.1 — property tracker landing zone (M5). Scraper appends new listings;
# `status` is human-edited (new/seen/contacted/dismissed). Append-only via
# append_rows(); dedup on listing_id (the scraper's seen.json + read_column).
PROPERTY_LISTINGS_COLUMNS = [
    "listing_id", "portal", "first_seen", "price_ils", "rooms",
    "size_sqm", "location", "url", "status",
]


class SchemaDriftError(RuntimeError):
    """Header row disagrees with SPEC §6.1 — abort the run (§7.1)."""


# ---------------------------------------------------------------------------
# Dataclasses
# ---------------------------------------------------------------------------
@dataclass
class Reminder:
    row: int
    title: str
    domain: str
    owner: str
    due: date | None
    lead_times: list[int]
    recurrence: str
    status: str
    last_sent: date | None
    channel: str
    notes: str
    # Dashboard write-back columns M/N/O — blank on legacy sheets, that's fine.
    last_done_by: str = ""
    done_at: datetime | None = None
    write_queue_tombstone: datetime | None = None
    guide_url: str = ""


@dataclass
class CellWrite:
    """One cell of a Reminders write-back batch, addressed by field name so
    callers never hardcode column letters. Values: date → DD/MM/YYYY (§8.5),
    datetime → ISO-T text (round-trips exactly; Sheets won't coerce the T
    form into a locale-formatted date cell), None → clear."""
    row: int
    field: str
    value: object

    def __post_init__(self):
        if self.field not in FIELD_TO_COL:
            raise ValueError(f"unknown Reminders field {self.field!r}")
        if self.row < 2:
            raise ValueError("row 1 is the header — refusing to write it")


@dataclass
class Settings:
    """SPEC §6.4 Settings tab. usermap keys are lowercased emails."""
    usermap: dict[str, str] = field(default_factory=dict)
    lang: str = "he"

    def display_name(self, email: str, default: str = "") -> str:
        return self.usermap.get((email or "").strip().lower(), default)


# ---------------------------------------------------------------------------
# Value encoding (shared by both backends' write paths)
# ---------------------------------------------------------------------------
def encode_value(v) -> object:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.isoformat(timespec="seconds")
    if isinstance(v, date):
        return v.strftime("%d/%m/%Y")          # §6.1 col D contract
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    return v


# ---------------------------------------------------------------------------
# Backends — duck-typed: grid / batch_update / append / tabs
# ---------------------------------------------------------------------------
class XlsxBackend:
    """openpyxl against a local xlsx. Reads use data_only (cached formula
    values); writes reopen WITHOUT data_only so formulas survive a save."""

    def __init__(self, path: Path):
        self.path = Path(path)

    def tabs(self) -> list[str]:
        from openpyxl import load_workbook
        return load_workbook(self.path, read_only=True).sheetnames

    def grid(self, tab: str) -> list[list]:
        from openpyxl import load_workbook
        ws = load_workbook(self.path, data_only=True)[tab]
        return [[c.value for c in row] for row in ws.iter_rows()]

    def batch_update(self, tab: str, cells: Iterable[tuple[int, int, object]]) -> None:
        from openpyxl import load_workbook
        wb = load_workbook(self.path)            # data_only=False: keep formulas
        ws = wb[tab]
        for row, col, value in cells:
            ws.cell(row, col).value = None if value is None else value
        wb.save(self.path)

    def append(self, tab: str, header: list[str], rows: list[list]) -> None:
        from openpyxl import load_workbook
        wb = load_workbook(self.path)
        if tab in wb.sheetnames:
            ws = wb[tab]
            if ws.max_row == 1 and ws.cell(1, 1).value is None:
                ws.append(header)
        else:
            ws = wb.create_sheet(tab)
            ws.append(header)
        for r in rows:
            ws.append(r)
        wb.save(self.path)

    def delete_rows(self, tab: str, indices: list[int]) -> None:
        from openpyxl import load_workbook
        if not indices:
            return
        wb = load_workbook(self.path)
        if tab not in wb.sheetnames:
            return
        ws = wb[tab]
        for i in sorted(set(indices), reverse=True):  # high→low: indices stay valid
            ws.delete_rows(i, 1)
        wb.save(self.path)


class GSheetBackend:
    """The only gspread client (ENGINEERING §1). Small retry on quota/5xx.
    Reads: UNFORMATTED_VALUE + FORMATTED_STRING dates, so numbers come back
    typed and dates come back as the displayed DD/MM/YYYY strings the parsers
    already speak."""

    RETRY_STATUS = {429, 500, 502, 503}

    def __init__(self, sheet_id: str, sa_path: Path):
        import gspread  # imported here so the seed backend works without it
        if not sa_path.exists():
            raise RuntimeError(
                f"{config.SHEET_ID_ENV} is set but no service-account JSON at "
                f"{sa_path} — refusing to fall back to the seed xlsx (fail loud)")
        self._gspread = gspread
        gc = gspread.service_account(filename=str(sa_path))
        self.sh = self._retry(gc.open_by_key, sheet_id)

    def _retry(self, fn, *a, **kw):
        for attempt in range(3):
            try:
                return fn(*a, **kw)
            except self._gspread.exceptions.APIError as e:  # noqa: PERF203
                status = getattr(getattr(e, "response", None), "status_code", None)
                if status in self.RETRY_STATUS and attempt < 2:
                    wait = 2 ** (attempt + 1)
                    log.warning("Sheets API %s — retry in %ss", status, wait)
                    _time.sleep(wait)
                    continue
                raise

    def _ws(self, tab: str):
        return self._retry(self.sh.worksheet, tab)

    def tabs(self) -> list[str]:
        return [w.title for w in self._retry(self.sh.worksheets)]

    def grid(self, tab: str) -> list[list]:
        ws = self._ws(tab)
        return self._retry(
            ws.get_values,
            value_render_option="UNFORMATTED_VALUE",
            date_time_render_option="FORMATTED_STRING",
        )

    def batch_update(self, tab: str, cells: Iterable[tuple[int, int, object]]) -> None:
        from gspread.utils import rowcol_to_a1
        ws = self._ws(tab)
        data = [{"range": rowcol_to_a1(row, col), "values": [[encode_value(value)]]}
                for row, col, value in cells]
        if data:
            self._retry(ws.batch_update, data, value_input_option="USER_ENTERED")

    def append(self, tab: str, header: list[str], rows: list[list]) -> None:
        try:
            ws = self._ws(tab)
        except self._gspread.exceptions.WorksheetNotFound:
            ws = self._retry(self.sh.add_worksheet, title=tab,
                             rows=200, cols=max(len(header), 8))
            self._retry(ws.append_row, header, value_input_option="RAW")
        if rows:
            encoded = [[encode_value(v) for v in r] for r in rows]
            self._retry(ws.append_rows, encoded, value_input_option="RAW",
                        table_range="A1")

    def delete_rows(self, tab: str, indices: list[int]) -> None:
        if not indices:
            return
        try:
            ws = self._ws(tab)
        except self._gspread.exceptions.WorksheetNotFound:
            return
        # Delete contiguous runs high→low so an earlier delete never shifts the
        # index of one still to come (one API call per run, not per row).
        for start, end in reversed(_contiguous_runs(sorted(set(indices)))):
            self._retry(ws.delete_rows, start, end)


# ---------------------------------------------------------------------------
# Backend selection
# ---------------------------------------------------------------------------
def _contiguous_runs(sorted_indices: list[int]) -> list[tuple[int, int]]:
    """Collapse a sorted-ascending index list into inclusive (start, end) runs:
    [2,3,4,7,9,10] → [(2,4),(7,7),(9,10)]."""
    runs: list[list[int]] = []
    for i in sorted_indices:
        if runs and i == runs[-1][1] + 1:
            runs[-1][1] = i
        else:
            runs.append([i, i])
    return [(a, b) for a, b in runs]
_default_backend = None


def is_live() -> bool:
    """True when the live Google Sheet is configured. Write-back callers gate
    on this so a creds-less run never mutates the committed seed template."""
    config.load_env()
    return bool(os.environ.get(config.SHEET_ID_ENV, "").strip())


def backend(path: Optional[Path] = None):
    """Explicit path → that xlsx (tests, tooling). No path → the live Sheet
    when configured, else the seed xlsx."""
    global _default_backend
    if path is not None:
        return XlsxBackend(path)
    if _default_backend is None:
        if is_live():
            sa = Path(os.environ.get(config.SA_JSON_ENV, "") or config.SA_JSON_DEFAULT)
            _default_backend = GSheetBackend(
                os.environ[config.SHEET_ID_ENV].strip(), sa)
            log.info("sheet backend: live Google Sheet")
        else:
            _default_backend = XlsxBackend(config.SHEET_PATH)
    return _default_backend


def reset_backend() -> None:
    """Drop the cached default backend (tests / long-lived processes)."""
    global _default_backend
    _default_backend = None


# ---------------------------------------------------------------------------
# Schema-drift guard (SPEC §7.1)
# ---------------------------------------------------------------------------
def _norm(h) -> str:
    return str(h or "").strip().casefold()


def validate_reminders_header(header_row: list) -> list[str]:
    """Mismatch descriptions, [] when clean. Strict through col P; columns
    beyond P are tolerated (additive-only schema, ENGINEERING §9)."""
    problems = []
    for col, expected in REMINDERS_COLUMNS.items():
        got = header_row[col - 1] if col <= len(header_row) else None
        if _norm(got) != _norm(expected):
            problems.append(f"col {col}: expected {expected!r}, found {got!r}")
    return problems


def _schema_guard(grid: list[list], where: str) -> None:
    problems = validate_reminders_header(grid[0] if grid else [])
    if problems:
        config.LOGS_DIR.mkdir(parents=True, exist_ok=True)
        config.SCHEMA_DRIFT_FLAG.write_text(json.dumps({
            "at": datetime.now().isoformat(timespec="seconds"),
            "tab": config.REMINDERS_TAB, "where": where, "problems": problems,
        }, ensure_ascii=False, indent=1), encoding="utf-8")
        log.error("schema_drift (%s): %s", where, "; ".join(problems))
        raise SchemaDriftError(
            f"Reminders header drifted from SPEC §6.1 ({where}): "
            + "; ".join(problems))
    # Clean read heals the flag so the briefing stops warning.
    if config.SCHEMA_DRIFT_FLAG.exists():
        config.SCHEMA_DRIFT_FLAG.unlink(missing_ok=True)


def schema_drift_flag() -> Optional[dict]:
    """The pending drift report, or None. The weekly briefing surfaces this."""
    if not config.SCHEMA_DRIFT_FLAG.exists():
        return None
    try:
        return json.loads(config.SCHEMA_DRIFT_FLAG.read_text(encoding="utf-8"))
    except (ValueError, OSError):
        return {"problems": ["unreadable schema_drift.flag"]}


# ---------------------------------------------------------------------------
# Reminders — read
# ---------------------------------------------------------------------------
def parse_lead_times(raw) -> list[int]:
    """Column E: CSV of day offsets ('60,30,7,1'). Junk filtered, sorted
    descending; default [7, 1] when empty/unparseable."""
    if raw is None:
        return [7, 1]
    if isinstance(raw, (int, float)):
        return [int(raw)]
    parts = [p.strip() for p in str(raw).split(",") if p.strip()]
    out = []
    for p in parts:
        try:
            out.append(int(p))
        except ValueError:
            pass
    return sorted(out, reverse=True) or [7, 1]


def _cell(row: list, col: int):
    v = row[col - 1] if col <= len(row) else None
    return None if (isinstance(v, str) and not v.strip()) else v


def read_reminders(path: Optional[Path] = None) -> list[Reminder]:
    """Reminders tab → dataclasses, header-validated first. Blank titles and
    templated `[...]` rows are skipped; unparseable dates become None (caller
    classifies them out)."""
    grid = backend(path).grid(config.REMINDERS_TAB)
    _schema_guard(grid, where="read_reminders")
    out = []
    for idx, row in enumerate(grid[1:], start=2):
        title = _cell(row, 1)
        if not title or str(title).startswith("["):  # skip blanks / templated rows
            continue
        out.append(Reminder(
            row=idx,
            title=str(title),
            domain=str(_cell(row, 2) or "Other"),
            owner=str(_cell(row, 3) or "Adar"),
            due=to_date(_cell(row, 4)),
            lead_times=parse_lead_times(_cell(row, 5)),
            recurrence=str(_cell(row, 6) or "One-off"),
            status=str(_cell(row, 7) or "Pending"),
            last_sent=to_date(_cell(row, 8)),
            channel=str(_cell(row, 9) or "WhatsApp"),
            notes=str(_cell(row, 10) or "").strip(),
            last_done_by=str(_cell(row, 13) or "").strip(),
            done_at=to_datetime(_cell(row, 14)),
            write_queue_tombstone=to_datetime(_cell(row, 15)),
            guide_url=str(_cell(row, 16) or "").strip(),
        ))
    return out


# ---------------------------------------------------------------------------
# Reminders — write (engine write-backs land here; ONE batched call)
# ---------------------------------------------------------------------------
def update_reminders(writes: list[CellWrite], path: Optional[Path] = None) -> None:
    """Apply a write-back batch to the Reminders tab. Header-validated first —
    a drifted sheet must not be written by position (§7.1)."""
    if not writes:
        return
    b = backend(path)
    _schema_guard(b.grid(config.REMINDERS_TAB), where="update_reminders")
    b.batch_update(config.REMINDERS_TAB,
                   [(w.row, FIELD_TO_COL[w.field], w.value) for w in writes])
    log.info("Reminders write-back: %d cell(s) across %d row(s)",
             len(writes), len({w.row for w in writes}))


# ---------------------------------------------------------------------------
# Generic tab append (summarizer → WhatsApp_Inbox / WhatsApp_Archive)
# ---------------------------------------------------------------------------
def append_rows(tab: str, columns: list[str], rows: list[dict],
                path: Optional[Path] = None) -> None:
    """Append dict-rows in `columns` order; the tab (with header) is created
    when missing — additive-only, nothing existing is touched."""
    if not rows:
        return
    backend(path).append(tab, columns,
                         [[r.get(c, "") for c in columns] for r in rows])


def read_column(tab: str, column: str, path: Optional[Path] = None) -> list:
    """All values of one named column (header-matched, loose). Missing tab or
    column → [] — readers degrade, they don't crash."""
    b = backend(path)
    if tab not in b.tabs():
        return []
    grid = b.grid(tab)
    if not grid:
        return []
    try:
        i = [_norm(h) for h in grid[0]].index(_norm(column))
    except ValueError:
        return []
    return [row[i] for row in grid[1:] if i < len(row) and row[i] not in (None, "")]


def roll_off_old_rows(tab: str, date_column: str, cutoff: date,
                      path: Optional[Path] = None) -> int:
    """Delete rows of `tab` whose `date_column` parses to a date strictly before
    `cutoff` (SPEC §6.2 hot-tab rolloff — WhatsApp_Inbox keeps 30 days, Archive
    keeps text forever). The header is kept; rows whose date is blank or
    unparseable are KEPT (never delete what we can't date). Live backend or an
    explicit path only — refuses to mutate the committed seed, exactly like the
    write-backs. Returns the number of rows deleted."""
    if path is None and not is_live():
        log.info("roll_off skipped for %s — no live backend (won't touch the seed)", tab)
        return 0
    b = backend(path)
    if tab not in b.tabs():
        return 0
    grid = b.grid(tab)
    if len(grid) < 2:
        return 0
    try:
        ci = [_norm(h) for h in grid[0]].index(_norm(date_column))
    except ValueError:
        log.warning("roll_off: column %r absent in %s — nothing rolled off",
                    date_column, tab)
        return 0
    stale = []
    for idx, row in enumerate(grid[1:], start=2):
        d = to_date(row[ci] if ci < len(row) else None)
        if d is not None and d < cutoff:
            stale.append(idx)
    if stale:
        b.delete_rows(tab, stale)
        log.info("rolled off %d row(s) from %s older than %s", len(stale), tab, cutoff)
    return len(stale)


# ---------------------------------------------------------------------------
# Settings (SPEC §6.4): Key|Value rows — emails map to display names, plus lang
# ---------------------------------------------------------------------------
def read_settings(path: Optional[Path] = None) -> Settings:
    """Missing tab degrades to defaults (empty usermap, Hebrew) with a warning
    — briefings must still render on a sheet that predates the Settings tab."""
    b = backend(path)
    s = Settings()
    if config.SETTINGS_TAB not in b.tabs():
        log.warning("no %s tab — UserMap empty, lang=he", config.SETTINGS_TAB)
        return s
    for row in b.grid(config.SETTINGS_TAB)[1:]:
        key = str(_cell(row, 1) or "").strip()
        value = str(_cell(row, 2) or "").strip()
        if not key or not value:
            continue
        if "@" in key:
            s.usermap[key.lower()] = value
        elif key.casefold() == "lang":
            s.lang = value
    return s


# ---------------------------------------------------------------------------
# Loose read handle for multi-tab readers (weekly briefing)
# ---------------------------------------------------------------------------
class _Cell:
    __slots__ = ("value",)

    def __init__(self, value):
        self.value = value


class _Tab:
    """Quacks like an openpyxl worksheet for the read patterns the briefing
    uses: .max_row and .cell(r, c).value, 1-based."""

    def __init__(self, grid: list[list]):
        self._grid = grid

    @property
    def max_row(self) -> int:
        return max(len(self._grid), 1)

    def cell(self, row: int, col: int) -> _Cell:
        try:
            return _Cell(self._grid[row - 1][col - 1])
        except IndexError:
            return _Cell(None)


class _Workbook:
    def __init__(self, b):
        self._b = b
        self._cache: dict[str, _Tab] = {}

    @property
    def sheetnames(self) -> list[str]:
        return self._b.tabs()

    def __getitem__(self, tab: str) -> _Tab:
        if tab not in self._cache:
            self._cache[tab] = _Tab(self._b.grid(tab))
        return self._cache[tab]

    def __contains__(self, tab: str) -> bool:
        return tab in self.sheetnames


def workbook(path: Optional[Path] = None) -> _Workbook:
    """Read-only duck-typed workbook over the active backend. Tab grids are
    fetched lazily and cached for the call's lifetime."""
    return _Workbook(backend(path))


def load_workbook_data(path: Path):
    """Deprecated M1 name — openpyxl-shaped read handle. Kept callable for
    one milestone; new code uses workbook()."""
    return workbook(path)
