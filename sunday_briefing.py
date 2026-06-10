"""
Family inc. — Sunday Briefing (Phase 1, dry-run / email-fallback)

Generates the weekly cross-domain briefing every Sunday 18:00. Reads
Family_OS.xlsx and writes:

  /Briefings/{YYYY-MM-DD}_sunday_briefing.md

Sections (in order):
  1. Week ahead — calendar events for next 7 days
  2. Reminders firing this week — items hitting a lead time or due
  3. Overdue items that need attention
  4. Money — current-month budget utilization vs target
  5. Goals — status, with any 90-day milestones in <= 30 days flagged
  6. Data hygiene — rows that are missing or stale

When the Baileys bridge is paired (QR scan + recipients.json on host), send_to_whatsapp()
via wa_outbox.queue_message() swaps in for the file write.
When Google Calendar is connected, the scheduled task syncs events into
the Calendar-Events tab BEFORE running this script — so this engine stays
data-source-agnostic.

Run modes:
  python sunday_briefing.py             # today's date, writes file
  python sunday_briefing.py --dry-run   # print only
  python sunday_briefing.py --as-of 2026-05-31
"""
from __future__ import annotations
import argparse
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).parent
SHEET_PATH = ROOT / "Family_OS.xlsx"
BRIEFINGS_DIR = ROOT / "Briefings"

WEEK_AHEAD_DAYS = 7
GOAL_MILESTONE_FLAG_DAYS = 30   # flag goals whose target date is within 30 days
STALE_GOAL_UPDATE_DAYS = 21     # warn if goal Last Update older than 3 weeks

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def to_date(v):
    if v is None: return None
    if isinstance(v, date) and not isinstance(v, datetime): return v
    if isinstance(v, datetime): return v.date()
    if isinstance(v, str):
        try: return datetime.strptime(v.strip(), "%Y-%m-%d").date()
        except ValueError: return None
    return None

def fmt_date(d: date) -> str:
    if not d: return ""
    return d.strftime("%a %b %-d")  # Sun May 31

def fmt_money(n) -> str:
    if n is None: return "—"
    try:
        n = float(n)
    except (TypeError, ValueError):
        return str(n)
    sign = "-" if n < 0 else ""
    return f"{sign}₪{abs(n):,.0f}"

def pct(num, denom):
    if not denom: return None
    return num / denom

# ---------------------------------------------------------------------------
# Section builders
# ---------------------------------------------------------------------------
def section_week_ahead(wb, today: date, end: date) -> str:
    ws = wb["Calendar-Events"]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = to_date(ws.cell(r, 1).value)
        if not d or d < today or d > end: continue
        title = ws.cell(r, 4).value
        if not title: continue
        rows.append({
            "date": d,
            "start": ws.cell(r, 2).value,
            "end": ws.cell(r, 3).value,
            "title": title,
            "owner": ws.cell(r, 5).value or "",
            "location": ws.cell(r, 7).value or "",
        })
    rows.sort(key=lambda x: (x["date"], x["start"] or ""))
    if not rows:
        return "_Quiet week ahead — no events scheduled._"
    by_day = {}
    for r in rows:
        by_day.setdefault(r["date"], []).append(r)
    lines = []
    for d in sorted(by_day):
        marker = " **(today)**" if d == today else ""
        lines.append(f"\n**{fmt_date(d)}**{marker}")
        for r in by_day[d]:
            time = ""
            if r["start"]: time = f"{r['start']}"
            if r["start"] and r["end"]: time += f"–{r['end']}"
            if not time: time = "all day"
            who = f" · {r['owner']}" if r["owner"] else ""
            loc = f" · {r['location']}" if r["location"] else ""
            lines.append(f"- {time} — {r['title']}{who}{loc}")
    return "\n".join(lines).lstrip()

def section_reminders_week(wb, today: date, end: date) -> tuple[str, str]:
    """Returns (upcoming_section, overdue_section)."""
    ws = wb["Reminders"]
    upcoming, overdue = [], []
    for r in range(2, ws.max_row + 1):
        title = ws.cell(r, 1).value
        if not title or str(title).startswith("["): continue
        status = ws.cell(r, 7).value or "Pending"
        if status in {"Done", "Skipped"}: continue
        due = to_date(ws.cell(r, 4).value)
        if not due: continue
        owner = ws.cell(r, 3).value or ""
        domain = ws.cell(r, 2).value or ""
        days = (due - today).days
        if days < 0:
            overdue.append((due, days, title, owner, domain))
        elif days <= WEEK_AHEAD_DAYS:
            upcoming.append((due, days, title, owner, domain))
    upcoming.sort(key=lambda x: x[0])
    overdue.sort(key=lambda x: x[0])

    def render(rows, kind: str) -> str:
        if not rows:
            return f"_No {kind} items._"
        lines = []
        for due, days, title, owner, domain in rows:
            if days < 0:
                tag = f"🔴 overdue {-days}d"
            elif days == 0:
                tag = "🟠 today"
            elif days <= 1:
                tag = "🟠 tomorrow"
            elif days <= 7:
                tag = f"🟡 in {days}d"
            else:
                tag = f"in {days}d"
            who = f" · {owner}" if owner else ""
            dom = f" [{domain}]" if domain else ""
            lines.append(f"- {tag} — {title}{who}{dom}  ({due.isoformat()})")
        return "\n".join(lines)

    return render(upcoming, "upcoming"), render(overdue, "overdue")

def section_money(wb, today: date) -> str:
    """Read Finance-Bdgt; we use the as-of (I1) which should be today."""
    ws = wb["Finance-Bdgt"]
    rows = []
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(r, 1).value
        if not cat or cat == "TOTAL": continue
        target = ws.cell(r, 2).value
        actual = ws.cell(r, 3).value
        if target in (None, 0): continue
        rows.append({
            "cat": cat,
            "target": float(target),
            "actual": float(actual or 0),
        })
    # Show top 3 over-budget and totals
    over = sorted([r for r in rows if r["actual"] > r["target"]],
                  key=lambda r: r["actual"] - r["target"], reverse=True)[:3]
    total_target = sum(r["target"] for r in rows)
    total_actual = sum(r["actual"] for r in rows)
    p = pct(total_actual, total_target)
    lines = [f"**Month-to-date:** {fmt_money(total_actual)} of {fmt_money(total_target)} target  ({(p or 0)*100:.0f}%)"]
    if over:
        lines.append("\n**Over-budget categories:**")
        for r in over:
            o = r["actual"] - r["target"]
            lines.append(f"- {r['cat']}: {fmt_money(r['actual'])} / {fmt_money(r['target'])}  (+{fmt_money(o)})")
    else:
        lines.append("\nNo categories over budget this month.")
    return "\n".join(lines)

def section_goals(wb, today: date) -> str:
    ws = wb["Goals"]
    rows = []
    for r in range(2, ws.max_row + 1):
        goal = ws.cell(r, 1).value
        if not goal: continue
        target = to_date(ws.cell(r, 4).value)
        milestone = ws.cell(r, 5).value or ""
        pct_done = ws.cell(r, 6).value
        last_update = to_date(ws.cell(r, 7).value)
        status = ws.cell(r, 8).value or "Not started"
        owner = ws.cell(r, 2).value or ""
        rows.append({
            "goal": goal, "owner": owner, "target": target, "milestone": milestone,
            "pct": pct_done, "last_update": last_update, "status": status
        })
    if not rows:
        return "_No goals tracked yet._"
    lines = []
    for r in rows:
        flags = []
        if r["target"] and (r["target"] - today).days <= GOAL_MILESTONE_FLAG_DAYS and (r["target"] - today).days >= 0:
            flags.append(f"⏰ milestone in {(r['target']-today).days}d")
        if r["last_update"] and (today - r["last_update"]).days > STALE_GOAL_UPDATE_DAYS:
            flags.append(f"⚠️ no update in {(today - r['last_update']).days}d")
        pct_str = f"{int(r['pct'])}%" if isinstance(r["pct"], (int,float)) else "—"
        flag_str = ("  " + " · ".join(flags)) if flags else ""
        line = f"- **{r['goal']}** ({r['owner']}, {r['status']}, {pct_str})"
        if r["milestone"]:
            line += f"\n  next: {r['milestone']}"
        line += flag_str
        lines.append(line)
    return "\n".join(lines)

def section_hygiene(wb, today: date) -> str:
    issues = []
    # Reminders missing due date
    r_ws = wb["Reminders"]
    missing = 0
    for row in range(2, r_ws.max_row + 1):
        if r_ws.cell(row, 1).value and not r_ws.cell(row, 4).value:
            missing += 1
    if missing: issues.append(f"- {missing} reminder(s) missing a Due Date")
    # Stale Last Imported on accounts
    fa_ws = wb["Finance-Accts"]
    for row in range(2, fa_ws.max_row + 1):
        if fa_ws.cell(row, 1).value and not fa_ws.cell(row, 1).value.startswith("["):
            last_imp = to_date(fa_ws.cell(row, 7).value)
            if not last_imp:
                issues.append(f"- Account `{fa_ws.cell(row, 1).value}` has no Last Imported date")
            elif (today - last_imp).days > 35:
                issues.append(f"- Account `{fa_ws.cell(row, 1).value}` not imported in {(today-last_imp).days}d")
    # Placeholder rows still in People/Health/etc.
    p_ws = wb["People"]
    placeholders = sum(1 for r in range(2, p_ws.max_row + 1)
                       if p_ws.cell(r, 1).value and str(p_ws.cell(r, 1).value).startswith("["))
    if placeholders: issues.append(f"- {placeholders} People row(s) still using placeholder names")
    if not issues:
        return "_All clean._"
    return "\n".join(issues)

# ---------------------------------------------------------------------------
# Render
# ---------------------------------------------------------------------------
def render_briefing(wb, today: date) -> str:
    end = today + timedelta(days=WEEK_AHEAD_DAYS)
    parts = []
    parts.append(f"# 🏠 Family inc. — Sunday Briefing")
    parts.append(f"_{today.strftime('%A, %B %-d, %Y')}_  ·  week of {today.isoformat()} → {end.isoformat()}\n")

    parts.append("## Week ahead")
    parts.append(section_week_ahead(wb, today, end))

    upcoming, overdue = section_reminders_week(wb, today, end)
    parts.append("\n## Reminders firing this week")
    parts.append(upcoming)
    parts.append("\n## Overdue")
    parts.append(overdue)

    parts.append("\n## Money")
    parts.append(section_money(wb, today))

    parts.append("\n## Goals")
    parts.append(section_goals(wb, today))

    parts.append("\n## Data hygiene")
    parts.append(section_hygiene(wb, today))

    parts.append("\n---\n_Read together with coffee, ~20 minutes. Edits go into Family_OS.xlsx — next week's briefing reflects them automatically._")
    return "\n".join(parts) + "\n"

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def run(today: date, dry_run: bool = False) -> Path | None:
    wb = load_workbook(SHEET_PATH, data_only=True)
    body = render_briefing(wb, today)
    if dry_run:
        print(body)
        return None
    BRIEFINGS_DIR.mkdir(exist_ok=True)
    out = BRIEFINGS_DIR / f"{today.isoformat()}_sunday_briefing.md"
    out.write_text(body, encoding="utf-8")
    print(f"wrote {out}")
    return out

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--as-of", help="YYYY-MM-DD; defaults to today")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()
    today = datetime.strptime(args.as_of, "%Y-%m-%d").date() if args.as_of else date.today()
    run(today, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
